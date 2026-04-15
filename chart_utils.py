"""
chart_utils.py — Plotly 차트 데이터 XLSX 추출 유틸리티

모든 분석 모듈에서 공용으로 사용합니다.
st.plotly_chart() 직후에 chart_download_btn()을 호출하면
차트 데이터와 차트 이미지가 포함된 XLSX 파일을 다운로드할 수 있는 버튼이 표시됩니다.

차트 이미지는 kaleido(Chrome 필요) 대신 matplotlib으로 재렌더링합니다.
"""
from __future__ import annotations
from typing import Optional, Tuple
import io
import pandas as pd


# ──────────────────────────────────────────────────────────────
# 데이터 추출
# ──────────────────────────────────────────────────────────────

def fig_to_df(fig) -> Optional[pd.DataFrame]:
    """Plotly figure의 모든 trace 데이터를 DataFrame으로 변환합니다."""
    try:
        rows = []
        for trace in fig.data:
            name = getattr(trace, 'name', '') or ''
            trace_type = getattr(trace, 'type', '') or ''

            if trace_type == 'heatmap' and getattr(trace, 'z', None) is not None:
                z = trace.z
                x_lbls = list(trace.x) if trace.x is not None else list(range(len(z[0]) if z else 0))
                y_lbls = list(trace.y) if trace.y is not None else list(range(len(z)))
                for yi, row_vals in enumerate(z):
                    for xi, val in enumerate(row_vals):
                        rows.append({'Y축': y_lbls[yi], 'X축': x_lbls[xi], '값': val, '계열': name})

            elif getattr(trace, 'r', None) is not None and getattr(trace, 'theta', None) is not None:
                for r_val, t_val in zip(trace.r, trace.theta):
                    rows.append({'방향': t_val, '빈도/값': r_val, '계열': name})

            elif trace_type == 'box':
                y_vals = getattr(trace, 'y', None)
                x_vals = getattr(trace, 'x', None)
                if y_vals is not None:
                    for i, yv in enumerate(y_vals):
                        row = {'값': yv, '계열': name}
                        if x_vals is not None:
                            row['X'] = x_vals[i] if i < len(x_vals) else ''
                        rows.append(row)

            elif trace_type == 'histogram':
                x_vals = getattr(trace, 'x', None)
                if x_vals is not None:
                    for xv in x_vals:
                        rows.append({'값': xv, '계열': name})

            elif getattr(trace, 'x', None) is not None and getattr(trace, 'y', None) is not None:
                for xv, yv in zip(trace.x, trace.y):
                    rows.append({'X': xv, 'Y': yv, '계열': name})

        return pd.DataFrame(rows) if rows else None
    except Exception:
        return None


# ──────────────────────────────────────────────────────────────
# matplotlib 한글 폰트 설정 (최초 1회)
# ──────────────────────────────────────────────────────────────

_mpl_font_configured = False

def _configure_mpl_font() -> None:
    """Streamlit Cloud의 NanumGothic 폰트를 matplotlib에 등록합니다."""
    global _mpl_font_configured
    if _mpl_font_configured:
        return
    try:
        import matplotlib.font_manager as fm
        import matplotlib.pyplot as plt

        # 시스템 폰트 중 Nanum 계열 탐색
        nanum_paths = [f for f in fm.findSystemFonts(fontpaths=None)
                       if 'nanum' in f.lower()]
        if nanum_paths:
            # NanumGothic 우선, 없으면 첫 번째 Nanum 폰트 사용
            gothic = [f for f in nanum_paths if 'gothic' in f.lower() and 'bold' not in f.lower()]
            chosen = gothic[0] if gothic else nanum_paths[0]
            fe = fm.FontEntry(fname=chosen, name='NanumGothic')
            fm.fontManager.ttflist.append(fe)
            plt.rcParams['font.family'] = 'NanumGothic'
        else:
            plt.rcParams['font.family'] = 'DejaVu Sans'

        plt.rcParams['axes.unicode_minus'] = False
    except Exception:
        pass
    _mpl_font_configured = True


# ──────────────────────────────────────────────────────────────
# matplotlib 차트 렌더링
# ──────────────────────────────────────────────────────────────

def _fig_to_png_matplotlib(fig) -> Tuple[Optional[bytes], Optional[str]]:
    """Plotly figure의 데이터를 matplotlib으로 재렌더링하여 PNG bytes를 반환합니다.

    Returns:
        (png_bytes, error_msg)
    """
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt

        _configure_mpl_font()

        df = fig_to_df(fig)
        if df is None or df.empty:
            return None, "추출 가능한 데이터가 없습니다."

        # 차트 제목 추출
        title = ''
        try:
            title = fig.layout.title.text or ''
        except Exception:
            pass

        cols = set(df.columns)
        fig_mpl, ax = plt.subplots(figsize=(12, 5))

        # ── X·Y 계열 (선/막대/산점도) ────────────────────────
        if 'X' in cols and 'Y' in cols:
            if '계열' in cols:
                for name, grp in df.groupby('계열', sort=False):
                    ax.plot(grp['X'], grp['Y'],
                            label=str(name), marker='o',
                            markersize=3, linewidth=1.5)
                if df['계열'].nunique() > 1:
                    ax.legend(fontsize=8, loc='best')
            else:
                ax.plot(df['X'], df['Y'], marker='o', markersize=3)
            ax.set_xlabel('X')
            ax.set_ylabel('Y')
            ax.tick_params(axis='x', rotation=45)

        # ── 극좌표 (바람장미) ─────────────────────────────────
        elif '방향' in cols and '빈도/값' in cols:
            if '계열' in cols:
                for name, grp in df.groupby('계열', sort=False):
                    ax.bar(range(len(grp)), grp['빈도/값'],
                           label=str(name), alpha=0.7)
                ax.set_xticks(range(len(df['방향'].unique())))
                ax.set_xticklabels(df['방향'].unique(), rotation=45, fontsize=8)
                ax.legend(fontsize=8)
            else:
                ax.bar(df['방향'], df['빈도/값'])
                ax.tick_params(axis='x', rotation=45)
            ax.set_xlabel('방향')
            ax.set_ylabel('빈도/값')

        # ── 히트맵 ───────────────────────────────────────────
        elif 'Y축' in cols and 'X축' in cols and '값' in cols:
            pivot = df.pivot_table(index='Y축', columns='X축',
                                   values='값', aggfunc='mean')
            im = ax.imshow(pivot.values, aspect='auto', cmap='RdYlBu_r')
            ax.set_xticks(range(len(pivot.columns)))
            ax.set_xticklabels(pivot.columns, rotation=45,
                                ha='right', fontsize=7)
            ax.set_yticks(range(len(pivot.index)))
            ax.set_yticklabels(pivot.index, fontsize=7)
            fig_mpl.colorbar(im, ax=ax, shrink=0.8)

        # ── 박스플롯 / 히스토그램 ────────────────────────────
        elif '값' in cols:
            if '계열' in cols and df['계열'].nunique() > 1:
                grp_list = [(n, g) for n, g in df.groupby('계열', sort=False)]
                data = [g['값'].dropna().values for _, g in grp_list]
                labels = [str(n) for n, _ in grp_list]
                ax.boxplot(data, labels=labels, patch_artist=True)
            else:
                ax.hist(df['값'].dropna(), bins=30, color='steelblue', edgecolor='white')
            ax.set_ylabel('값')

        if title:
            ax.set_title(title, fontsize=11, fontweight='bold')

        plt.tight_layout()
        buf = io.BytesIO()
        fig_mpl.savefig(buf, format='png', dpi=150, bbox_inches='tight')
        plt.close(fig_mpl)
        buf.seek(0)
        return buf.read(), None

    except Exception as e:
        return None, f"{type(e).__name__}: {e}"


# ──────────────────────────────────────────────────────────────
# XLSX 생성
# ──────────────────────────────────────────────────────────────

def _build_xlsx(fig, filename: str) -> Tuple[Optional[bytes], Optional[str]]:
    """데이터 시트 + 차트 이미지 시트가 포함된 XLSX bytes를 반환합니다.

    Returns:
        (xlsx_bytes, chart_error_msg)
    """
    chart_error: Optional[str] = None
    try:
        import openpyxl
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ── 데이터 시트 ──────────────────────────────────────
        ws_data = wb.active
        ws_data.title = "데이터"

        df = fig_to_df(fig)
        if df is not None and not df.empty:
            for ci, col in enumerate(df.columns, 1):
                cell = ws_data.cell(row=1, column=ci, value=col)
                cell.font = openpyxl.styles.Font(bold=True)
            for ri, row_tuple in enumerate(df.itertuples(index=False), 2):
                for ci, val in enumerate(row_tuple, 1):
                    ws_data.cell(row=ri, column=ci, value=val)
            for ci, col in enumerate(df.columns, 1):
                max_len = max(len(str(col)), df[col].astype(str).str.len().max())
                ws_data.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 40)

        # ── 차트 이미지 시트 (matplotlib) ────────────────────
        png_bytes, chart_error = _fig_to_png_matplotlib(fig)
        if png_bytes is not None:
            ws_chart = wb.create_sheet("차트")
            img = XLImage(io.BytesIO(png_bytes))
            img.anchor = 'A1'
            ws_chart.add_image(img)

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue(), chart_error

    except Exception as e:
        return None, f"{type(e).__name__}: {e}"


# ──────────────────────────────────────────────────────────────
# 공개 API
# ──────────────────────────────────────────────────────────────

def chart_download_btn(fig, key: str, filename: str = "chart_data") -> None:
    """Plotly figure 데이터를 차트 이미지가 포함된 XLSX로 다운로드하는 버튼을 표시합니다.

    성능 최적화: XLSX(matplotlib PNG 포함)는 사용자가 버튼을 클릭할 때만 생성합니다.
    생성된 결과는 session_state에 캐시되어 이후 렌더링에서 재사용됩니다.

    - 데이터 시트: trace 데이터 (모든 차트 타입 지원)
    - 차트 시트: matplotlib으로 재렌더링한 PNG 이미지
    - 완전 실패 시 CSV로 폴백

    Args:
        fig:      st.plotly_chart()에 넘긴 것과 동일한 Plotly figure 객체
        key:      Streamlit 위젯 고유 키 (앱 전체에서 중복 불가)
        filename: 다운로드 파일명 (.xlsx 확장자 자동 추가)
    """
    import streamlit as st

    cache_key = f"__xlsx_cache__{key}"

    # ── 이미 생성된 경우 → 다운로드 버튼 바로 표시 ──────────────
    if cache_key in st.session_state:
        xlsx_bytes, chart_error = st.session_state[cache_key]
        if xlsx_bytes is not None:
            st.download_button(
                "⬇️ 차트 XLSX 다운로드",
                xlsx_bytes,
                f"{filename}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=key,
            )
        else:
            # 폴백: CSV
            df = fig_to_df(fig)
            if df is not None and not df.empty:
                csv = df.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')
                st.download_button(
                    "⬇️ 차트 데이터 CSV",
                    csv,
                    f"{filename}.csv",
                    "text/csv",
                    key=key,
                )
        if chart_error:
            st.caption(f"⚠️ 차트 이미지 생성 실패 (데이터 시트만 포함): {chart_error}")
        return

    # ── 미생성 상태 → 클릭 시 생성 트리거 ────────────────────────
    if st.button("⬇️ 차트 XLSX 다운로드", key=f"{key}__prepare"):
        with st.spinner("XLSX 생성 중..."):
            xlsx_bytes, chart_error = _build_xlsx(fig, filename)
        st.session_state[cache_key] = (xlsx_bytes, chart_error)
        st.rerun()
