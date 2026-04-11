# excel_generator.py — 엑셀 보고서 생성 모듈 (v5.0)

import io
import io as _io
from datetime import datetime

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image as XLImage

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 색상 팔레트 상수
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

C = {
    'dark_blue':    '1F4E79',
    'mid_blue':     '2E75B6',
    'light_blue':   'DEEAF1',
    'white':        'FFFFFF',
    'orange':       'C55A11',
    'light_orange': 'FCE4D6',
    'green':        '375623',
    'light_green':  'E2EFDA',
    'mid_green':    '70AD47',
    'gray':         '595959',
    'light_gray':   'F2F2F2',
    'border':       'BFBFBF',
    'yellow':       'FFE699',
    'light_yellow': 'FFF2CC',
}

FONT = '맑은 고딕'
RAW_SHEET = "'📊 원본 데이터'"

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SUM/AVG 구분
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

SUM_ELEMENTS = {
    'precipitation', 'sunshine', 'solar_rad', 'snowfall',
    'wind_run_100m', 'daylight_hours', 'evaporation_large', 'evaporation_small',
}

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 요소 레이블 매핑
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

ELEMENT_LABELS = {
    'temp_avg':        '평균기온(℃)',
    'temp_max':        '최고기온(℃)',
    'temp_min':        '최저기온(℃)',
    'dew_point':       '이슬점온도(℃)',
    'frost_temp':      '초상온도(℃)',
    'precipitation':   '강수량(mm)',
    'precip_1hr_max':  '1시간최다강수(mm)',
    'humidity':        '습도(%)',
    'humidity_min':    '최소습도(%)',
    'vapor_pressure':  '증기압(hPa)',
    'wind_speed':      '평균풍속(m/s)',
    'wind_max':        '최대풍속(m/s)',
    'wind_gust':       '최대순간풍속(m/s)',
    'wind_run_100m':   '풍정합100m(m)',
    'pressure_sea':    '해면기압(hPa)',
    'pressure_local':  '현지기압(hPa)',
    'sunshine':        '일조시간(hr)',
    'daylight_hours':  '가조시간(hr)',
    'solar_rad':       '일사량(MJ/m²)',
    'cloud_cover':     '전운량(1/10)',
    'snowfall':        '신적설(cm)',
    'snow_depth':      '적설깊이(cm)',
    'evaporation_large':  '대형증발량(mm)',
    'evaporation_small':  '소형증발량(mm)',
    'soil_temp_surface':  '지면온도(℃)',
    'soil_temp_5cm':   '5cm지중온도(℃)',
    'soil_temp_10cm':  '10cm지중온도(℃)',
    'soil_temp_20cm':  '20cm지중온도(℃)',
    'soil_temp_30cm':  '30cm지중온도(℃)',
    'soil_temp_50cm':  '0.5m지중온도(℃)',
    'soil_temp_100cm': '1.0m지중온도(℃)',
    'soil_temp_150cm': '1.5m지중온도(℃)',
    'soil_temp_300cm': '3.0m지중온도(℃)',
    'soil_temp_500cm': '5.0m지중온도(℃)',
    'fog_duration':    '안개시간(hr)',
}

# 기상 특성 검토 대상 요소
CHAR_ELEMS = [
    ('temp_avg',      '평균기온(℃)',     'AVERAGEIFS', '0.0'),
    ('temp_max',      '최고기온(℃)',     'AVERAGEIFS', '0.0'),
    ('temp_min',      '최저기온(℃)',     'AVERAGEIFS', '0.0'),
    ('precipitation', '강수량(mm)',      'SUMIFS_AVG', '#,##0.0'),
    ('humidity',      '습도(%)',         'AVERAGEIFS', '0.0'),
    ('wind_speed',    '풍속(m/s)',       'AVERAGEIFS', '0.0'),
    ('sunshine',      '일조시간(hr)',    'SUMIFS_AVG', '#,##0.0'),
    ('solar_rad',     '일사량(MJ/m²)',   'SUMIFS_AVG', '#,##0.0'),
]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 헬퍼 함수
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _thin(color=None):
    clr = color if color else C['border']
    s = Side(style='thin', color=clr)
    return Border(left=s, right=s, top=s, bottom=s)


def _hc(ws, row, col, val, *, bg=None, fg=None,
        bold=True, sz=10, align='center', wrap=False):
    bg = bg if bg else C['dark_blue']
    fg = fg if fg else C['white']
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name=FONT, bold=bold, color=fg, size=sz)
    c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    c.border = _thin()
    return c


def _dc(ws, row, col, val, *, bg=None, bold=False, sz=9,
        align='center', nf=None, wrap=False):
    bg = bg if bg else C['white']
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name=FONT, bold=bold, size=sz, color='000000')
    c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    c.border = _thin()
    if nf:
        c.number_format = nf
    return c


def _title(ws, row, cs, ce, text, h=24, bg=None, sz=12):
    bg = bg if bg else C['dark_blue']
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    c = ws.cell(row=row, column=cs, value=text)
    c.font = Font(name=FONT, bold=True, size=sz, color=C['white'])
    c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = h
    return c


def _note(ws, row, cs, ce, text):
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    c = ws.cell(row=row, column=cs, value=text)
    c.font = Font(name=FONT, size=8, color=C['gray'])
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 13
    return c


def _safe_val(v):
    """NaN/inf 를 None 으로 변환"""
    if v is None:
        return None
    try:
        if np.isnan(v) or np.isinf(v):
            return None
    except (TypeError, ValueError):
        pass
    return v


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ExcelReportGenerator
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class ExcelReportGenerator:
    """기상자료 Excel 보고서 생성기"""

    def __init__(self):
        self._col = {}      # 요소키 → Excel 컬럼 레터 (원본데이터 기준)
        self._years = []
        self._n_rows = 0
        self._has_stn = False

    # ──────────────────────────────────
    # 공개 인터페이스
    # ──────────────────────────────────

    def generate_excel(self, df, monthly_df=None, climate_df=None) -> bytes:
        if df is None or df.empty:
            raise ValueError("데이터가 없습니다.")

        df2 = df.copy()
        if 'date' in df2.columns:
            df2['date'] = pd.to_datetime(df2['date'])
        if 'year' not in df2.columns and 'date' in df2.columns:
            df2['year'] = df2['date'].dt.year
        if 'month' not in df2.columns and 'date' in df2.columns:
            df2['month'] = df2['date'].dt.month

        self._years = sorted(df2['year'].dropna().unique().astype(int).tolist())
        self._n_rows = len(df2)
        self._has_stn = 'station_name' in df2.columns
        self._raw2_total_rows = 1
        self._raw2_item_labels = []
        self._raw2_cache = None

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        self._sheet_summary(wb, df2)
        self._sheet_raw(wb, df2)           # 반드시 두 번째 (_col 세팅)
        self._sheet_weather_chars(wb, df2)
        self._sheet_climate_change(wb, df2)
        self._sheet_precipitation(wb, df2)
        self._sheet_cumulative_precip(wb, df2)
        self._sheet_rainfall_days(wb, df2)
        self._sheet_pivot(wb, df2)
        self._sheet_weather_overview(wb, df2)
        self._sheet_climate_monthly(wb, df2)
        try:
            self._sheet_raw2(wb, df2)
        except Exception:
            pass
        try:
            self._sheet_pivot_work(wb, df2)
        except Exception:
            pass
        try:
            self._sheet_pivot_work2(wb, df2)
        except Exception:
            pass
        try:
            self._sheet_boxplot(wb, df2)
        except Exception:
            pass

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        raw_bytes = output.getvalue()
        # 피벗 테이블 XML 직접 주입
        try:
            return self._inject_pivot_tables(raw_bytes, df2)
        except Exception:
            return raw_bytes

    def generate_filename(self, prefix="기상보고서") -> str:
        today = datetime.today().strftime("%Y%m%d")
        return f"{prefix}_{today}.xlsx"

    # ──────────────────────────────────
    # 1. 보고서 요약 시트
    # ──────────────────────────────────

    def _sheet_summary(self, wb, df):
        ws = wb.create_sheet("📋 보고서 요약")
        ws.sheet_view.showGridLines = False

        # 컬럼 폭
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 16

        # 메인 타이틀
        _title(ws, 1, 2, 6, '기상자료 현황 보고서', h=36, sz=16)

        # 분석 정보
        yr_min = min(self._years) if self._years else '-'
        yr_max = max(self._years) if self._years else '-'
        today_str = datetime.today().strftime('%Y년 %m월 %d일')

        r = 3
        _hc(ws, r, 2, '분석 기간', bg=C['mid_blue'], sz=9)
        _dc(ws, r, 3, f'{yr_min}년 ~ {yr_max}년', bg=C['light_blue'], align='center')
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        _hc(ws, r, 5, '작성일', bg=C['mid_blue'], sz=9)
        _dc(ws, r, 6, today_str, bg=C['light_blue'], align='center')

        r = 5
        _title(ws, r, 2, 6, '▶ 관측소별 주요 기상요소 요약', h=20, bg=C['mid_blue'], sz=10)

        # 헤더
        r = 6
        headers = ['관측소명', '평균기온(℃)', '연강수량(mm)', '평균풍속(m/s)', '일조시간(hr)']
        for ci, h in enumerate(headers, 2):
            _hc(ws, r, ci, h, sz=9)

        # 관측소별 데이터
        stations = df['station_name'].unique() if self._has_stn else ['전체']
        r = 7
        for stn in stations:
            sdf = df[df['station_name'] == stn] if self._has_stn else df
            _dc(ws, r, 2, stn if self._has_stn else '전체', bg=C['light_blue'], bold=True)

            def _avg(key):
                if key in sdf.columns:
                    v = sdf[key].mean()
                    return round(float(v), 1) if pd.notna(v) else None
                return None

            def _sum_yr(key):
                if key in sdf.columns and 'year' in sdf.columns:
                    yr_count = sdf['year'].nunique()
                    if yr_count > 0:
                        v = sdf[key].sum() / yr_count
                        return round(float(v), 1) if pd.notna(v) else None
                return None

            _dc(ws, r, 3, _avg('temp_avg'), nf='0.0')
            _dc(ws, r, 4, _sum_yr('precipitation'), nf='#,##0.0')
            _dc(ws, r, 5, _avg('wind_speed'), nf='0.0')
            _dc(ws, r, 6, _sum_yr('sunshine'), nf='#,##0.0')
            r += 1

        # 주석
        _note(ws, r + 1, 2, 6, '※ 연강수량·일조시간은 연도 수로 나눈 연평균값입니다.')

    # ──────────────────────────────────
    # 2. 원본 데이터 시트 (_col 매핑 세팅)
    # ──────────────────────────────────

    def _sheet_raw(self, wb, df):
        ws = wb.create_sheet("📊 원본 데이터")
        ws.freeze_panes = 'A2'

        # 컬럼 순서 결정
        elem_keys = [k for k in ELEMENT_LABELS if k in df.columns]

        # 헤더 행 작성
        col = 1
        _hc(ws, 1, col, '날짜');      ws.column_dimensions[get_column_letter(col)].width = 12
        col += 1
        _hc(ws, 1, col, '연도');      ws.column_dimensions[get_column_letter(col)].width = 7
        self._col['year'] = get_column_letter(col)
        col += 1
        _hc(ws, 1, col, '월');         ws.column_dimensions[get_column_letter(col)].width = 5
        self._col['month'] = get_column_letter(col)
        col += 1

        if self._has_stn:
            _hc(ws, 1, col, '관측소명'); ws.column_dimensions[get_column_letter(col)].width = 12
            self._col['station'] = get_column_letter(col)
            col += 1

        for key in elem_keys:
            label = ELEMENT_LABELS[key]
            _hc(ws, 1, col, label, wrap=True)
            ws.column_dimensions[get_column_letter(col)].width = max(len(label) * 1.1, 10)
            self._col[key] = get_column_letter(col)
            col += 1

        # 데이터 행 작성
        for ri, (_, row) in enumerate(df.iterrows(), 2):
            c = 1
            # 날짜
            date_val = row.get('date')
            if pd.notna(date_val):
                date_str = pd.to_datetime(date_val).strftime('%Y-%m-%d')
            else:
                date_str = ''
            _dc(ws, ri, c, date_str, nf='@'); c += 1

            # 연도 (수식)
            if date_str:
                ws.cell(row=ri, column=c, value=f'=YEAR(A{ri})')
            else:
                ws.cell(row=ri, column=c, value=None)
            ws.cell(row=ri, column=c).font = Font(name=FONT, size=9)
            ws.cell(row=ri, column=c).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=ri, column=c).border = _thin()
            ws.cell(row=ri, column=c).number_format = '0'
            c += 1

            # 월 (수식)
            if date_str:
                ws.cell(row=ri, column=c, value=f'=MONTH(A{ri})')
            else:
                ws.cell(row=ri, column=c, value=None)
            ws.cell(row=ri, column=c).font = Font(name=FONT, size=9)
            ws.cell(row=ri, column=c).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=ri, column=c).border = _thin()
            ws.cell(row=ri, column=c).number_format = '0'
            c += 1

            if self._has_stn:
                _dc(ws, ri, c, row.get('station_name', '')); c += 1

            for key in elem_keys:
                v = _safe_val(row.get(key))
                _dc(ws, ri, c, v, nf='#,##0.0')
                c += 1

        # Excel Table 생성
        if self._n_rows > 0:
            total_cols = col - 1
            ref = f"A1:{get_column_letter(total_cols)}{self._n_rows + 1}"
            try:
                tbl = Table(displayName="기상데이터", ref=ref)
                style = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                tbl.tableStyleInfo = style
                ws.add_table(tbl)
            except Exception:
                pass

        ws.row_dimensions[1].height = 30

    # ──────────────────────────────────
    # 3. 기상 특성 검토 시트
    # ──────────────────────────────────

    def _sheet_weather_chars(self, wb, df):
        ws = wb.create_sheet("🌡️ 기상 특성 검토")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'B5'

        avail = [(k, lbl, method, nf) for k, lbl, method, nf in CHAR_ELEMS if k in self._col]
        if not avail:
            _title(ws, 1, 1, 4, '기상 특성 검토 — 데이터 없음')
            return

        n_yr = len(self._years) if self._years else 1
        mc = self._col.get('month', 'C')

        # 컬럼 폭
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 10

        r = 1
        _title(ws, r, 1, len(avail) + 2, '기상 특성 검토')

        # ── 표 1: 월별 기후 특성 ──────────────────
        r = 2
        _title(ws, r, 1, len(avail) + 2, '표 1. 월별 기후 특성 (월평균/월합산)', h=20, bg=C['mid_blue'], sz=10)

        r = 3
        _hc(ws, r, 1, '구분')
        _hc(ws, r, 2, '월')
        for ci, (k, lbl, method, nf) in enumerate(avail, 3):
            _hc(ws, r, ci, lbl, wrap=True)
            ws.column_dimensions[get_column_letter(ci)].width = 13

        month_rows = {}  # month → row number (표1 기준)
        for m in range(1, 13):
            r += 1
            month_rows[m] = r
            _dc(ws, r, 1, f'{m}월', bg=C['light_blue'], bold=True)
            _dc(ws, r, 2, m, bg=C['light_blue'])
            for ci, (k, lbl, method, nf) in enumerate(avail, 3):
                ec = self._col[k]
                if method == 'AVERAGEIFS':
                    formula = f'=ROUND(AVERAGEIFS({RAW_SHEET}!{ec}:{ec},{RAW_SHEET}!{mc}:{mc},{m}),1)'
                else:
                    formula = f'=ROUND(SUMIFS({RAW_SHEET}!{ec}:{ec},{RAW_SHEET}!{mc}:{mc},{m})/{n_yr},1)'
                c = ws.cell(row=r, column=ci, value=formula)
                c.font = Font(name=FONT, size=9)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = _thin()
                c.number_format = nf

        # 연간 합계/평균 행
        r += 1
        annual_row = r
        _dc(ws, r, 1, '연간', bg=C['yellow'], bold=True)
        _dc(ws, r, 2, '-', bg=C['yellow'])
        for ci, (k, lbl, method, nf) in enumerate(avail, 3):
            start_r = month_rows[1]
            end_r = month_rows[12]
            cl = get_column_letter(ci)
            if method == 'AVERAGEIFS':
                formula = f'=ROUND(AVERAGE({cl}{start_r}:{cl}{end_r}),1)'
            else:
                formula = f'=ROUND(SUM({cl}{start_r}:{cl}{end_r}),1)'
            c = ws.cell(row=r, column=ci, value=formula)
            c.font = Font(name=FONT, size=9, bold=True)
            c.fill = PatternFill(start_color=C['yellow'], end_color=C['yellow'], fill_type='solid')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = _thin()
            c.number_format = nf

        r += 2

        # ── 표 2: 계절별 기상 특성 ──────────────────
        _title(ws, r, 1, len(avail) + 2, '표 2. 계절별 기상 특성', h=20, bg=C['mid_blue'], sz=10)
        r += 1
        _hc(ws, r, 1, '계절')
        _hc(ws, r, 2, '월')
        for ci, (k, lbl, method, nf) in enumerate(avail, 3):
            _hc(ws, r, ci, lbl, wrap=True)

        seasons = [
            ('봄', [3, 4, 5]),
            ('여름', [6, 7, 8]),
            ('가을', [9, 10, 11]),
            ('겨울', [12, 1, 2]),
        ]
        season_bg = {
            '봄': C['light_green'], '여름': C['light_orange'],
            '가을': C['light_yellow'], '겨울': C['light_blue'],
        }

        for sname, months in seasons:
            r += 1
            bg = season_bg[sname]
            _dc(ws, r, 1, sname, bg=bg, bold=True)
            _dc(ws, r, 2, '+'.join(str(m) + '월' for m in months), bg=bg)
            for ci, (k, lbl, method, nf) in enumerate(avail, 3):
                cl = get_column_letter(ci)
                refs = ','.join(f'{cl}{month_rows[m]}' for m in months if m in month_rows)
                if method == 'AVERAGEIFS':
                    formula = f'=ROUND(AVERAGE({refs}),1)'
                else:
                    formula = f'=ROUND(SUM({refs}),1)'
                c = ws.cell(row=r, column=ci, value=formula)
                c.font = Font(name=FONT, size=9)
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = _thin()
                c.number_format = nf

        r += 2

        # ── 표 3: 극값 분석 ──────────────────
        _title(ws, r, 1, len(avail) + 2, '표 3. 극값 분석 (전체 기간)', h=20, bg=C['mid_blue'], sz=10)
        r += 1
        _hc(ws, r, 1, '구분')
        _hc(ws, r, 2, '항목')
        for ci, (k, lbl, method, nf) in enumerate(avail, 3):
            _hc(ws, r, ci, lbl, wrap=True)

        for stat_name, func in [('최대값', 'MAX'), ('최솟값', 'MIN')]:
            r += 1
            _dc(ws, r, 1, stat_name, bg=C['light_gray'], bold=True)
            _dc(ws, r, 2, func, bg=C['light_gray'])
            for ci, (k, lbl, method, nf) in enumerate(avail, 3):
                ec = self._col[k]
                formula = f'={func}({RAW_SHEET}!{ec}:{ec})'
                c = ws.cell(row=r, column=ci, value=formula)
                c.font = Font(name=FONT, size=9)
                c.fill = PatternFill(start_color=C['light_gray'], end_color=C['light_gray'], fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = _thin()
                c.number_format = nf

        # ── 차트: 월별 기온 + 강수량 ──────────────────
        r += 2
        self._add_weather_chart(ws, avail, month_rows, r)

    def _add_weather_chart(self, ws, avail, month_rows, chart_row):
        """월별 기온 꺾은선 + 강수량 막대 차트"""
        try:
            temp_cols = [ci + 3 for ci, (k, _, _, _) in enumerate(avail) if k in ('temp_avg', 'temp_max', 'temp_min')]
            precip_cols = [ci + 3 for ci, (k, _, _, _) in enumerate(avail) if k == 'precipitation']

            if not temp_cols and not precip_cols:
                return

            start_r = month_rows[1]
            end_r = month_rows[12]

            # 기온 꺾은선
            if temp_cols:
                lc = LineChart()
                lc.title = '월별 기온'
                lc.style = 10
                lc.y_axis.title = '기온(℃)'
                lc.x_axis.title = '월'
                lc.width = 15
                lc.height = 10
                for tc in temp_cols:
                    data = Reference(ws, min_col=tc, min_row=start_r - 1, max_row=end_r)
                    lc.add_data(data, titles_from_data=True)
                cats = Reference(ws, min_col=2, min_row=start_r, max_row=end_r)
                lc.set_categories(cats)
                ws.add_chart(lc, f'A{chart_row}')

            # 강수량 막대
            if precip_cols:
                bc = BarChart()
                bc.title = '월별 강수량'
                bc.style = 10
                bc.y_axis.title = '강수량(mm)'
                bc.x_axis.title = '월'
                bc.width = 15
                bc.height = 10
                for pc in precip_cols:
                    data = Reference(ws, min_col=pc, min_row=start_r - 1, max_row=end_r)
                    bc.add_data(data, titles_from_data=True)
                cats = Reference(ws, min_col=2, min_row=start_r, max_row=end_r)
                bc.set_categories(cats)
                offset_col = 10
                ws.add_chart(bc, f'{get_column_letter(offset_col)}{chart_row}')
        except Exception:
            pass

    # ──────────────────────────────────
    # 4. 기후변화 검토 시트
    # ──────────────────────────────────

    def _sheet_climate_change(self, wb, df):
        ws = wb.create_sheet("🌍 기후변화 검토")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'B5'

        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 8

        r = 1
        _title(ws, r, 1, 9, '기후변화 검토')

        yc = self._col.get('year', 'B')
        mc = self._col.get('month', 'C')

        stat_cols = [
            ('temp_avg',      '연평균기온(℃)',   'AVERAGEIFS', '0.0'),
            ('temp_max',      '최고기온(℃)',     'AVERAGEIFS', '0.0'),
            ('temp_min',      '최저기온(℃)',     'AVERAGEIFS', '0.0'),
            ('precipitation', '총강수량(mm)',    'SUMIFS',     '#,##0.0'),
            ('precipitation', '월평균강수량(mm)', 'SUMIFS_12',  '#,##0.0'),
            ('humidity',      '평균습도(%)',     'AVERAGEIFS', '0.0'),
        ]
        avail_sc = [(k, lbl, meth, nf) for k, lbl, meth, nf in stat_cols if k in self._col]

        # ── 표 1: 연도별 기상 통계 ──────────────────
        r = 2
        _title(ws, r, 1, len(avail_sc) + 2, '표 1. 연도별 기상 통계', h=20, bg=C['mid_blue'], sz=10)
        r = 3
        _hc(ws, r, 1, 'No.')
        _hc(ws, r, 2, '연도')
        for ci, (k, lbl, meth, nf) in enumerate(avail_sc, 3):
            _hc(ws, r, ci, lbl, wrap=True)
            ws.column_dimensions[get_column_letter(ci)].width = 14

        year_rows = {}
        for yi, yr in enumerate(self._years):
            r += 1
            year_rows[yr] = r
            bg = C['light_blue'] if yi % 2 == 0 else C['white']
            _dc(ws, r, 1, yi + 1, bg=bg)
            _dc(ws, r, 2, yr, bg=bg, bold=True)
            for ci, (k, lbl, meth, nf) in enumerate(avail_sc, 3):
                ec = self._col[k]
                if meth == 'AVERAGEIFS':
                    formula = f'=ROUND(AVERAGEIFS({RAW_SHEET}!{ec}:{ec},{RAW_SHEET}!{yc}:{yc},{yr}),1)'
                elif meth == 'SUMIFS':
                    formula = f'=ROUND(SUMIFS({RAW_SHEET}!{ec}:{ec},{RAW_SHEET}!{yc}:{yc},{yr}),1)'
                else:  # SUMIFS_12
                    formula = f'=ROUND(SUMIFS({RAW_SHEET}!{ec}:{ec},{RAW_SHEET}!{yc}:{yc},{yr})/12,1)'
                c = ws.cell(row=r, column=ci, value=formula)
                c.font = Font(name=FONT, size=9)
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = _thin()
                c.number_format = nf

        # 집계 행
        if year_rows:
            first_yr_r = year_rows[self._years[0]]
            last_yr_r = year_rows[self._years[-1]]
            for stat_name, bg, func in [
                ('전체평균', C['light_yellow'], 'AVERAGE'),
                ('최대', C['light_orange'], 'MAX'),
                ('최소', C['light_blue'], 'MIN'),
            ]:
                r += 1
                _dc(ws, r, 1, stat_name, bg=bg, bold=True)
                _dc(ws, r, 2, '-', bg=bg)
                for ci, (k, lbl, meth, nf) in enumerate(avail_sc, 3):
                    cl = get_column_letter(ci)
                    formula = f'=ROUND({func}({cl}{first_yr_r}:{cl}{last_yr_r}),1)'
                    c = ws.cell(row=r, column=ci, value=formula)
                    c.font = Font(name=FONT, size=9, bold=True)
                    c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border = _thin()
                    c.number_format = nf

        r += 2

        # ── 표 2: 전반기·후반기 비교 ──────────────────
        if len(self._years) >= 2:
            mid = len(self._years) // 2
            first_half = self._years[:mid]
            second_half = self._years[mid:]
            _title(ws, r, 1, len(avail_sc) + 2,
                   f'표 2. 전반기({first_half[0]}~{first_half[-1]}) · 후반기({second_half[0]}~{second_half[-1]}) 비교',
                   h=20, bg=C['mid_blue'], sz=10)
            r += 1
            _hc(ws, r, 1, '구분')
            _hc(ws, r, 2, '기간')
            for ci, (k, lbl, meth, nf) in enumerate(avail_sc, 3):
                _hc(ws, r, ci, lbl, wrap=True)

            for period_name, years_list in [
                (f'전반기\n({first_half[0]}~{first_half[-1]})', first_half),
                (f'후반기\n({second_half[0]}~{second_half[-1]})', second_half),
            ]:
                r += 1
                is_first = '전반기' in period_name
                bg = C['light_blue'] if is_first else C['light_orange']
                _dc(ws, r, 1, period_name, bg=bg, bold=True, wrap=True)
                _dc(ws, r, 2, f'{years_list[0]}~{years_list[-1]}', bg=bg)
                ws.row_dimensions[r].height = 28
                for ci, (k, lbl, meth, nf) in enumerate(avail_sc, 3):
                    yr_refs = [year_rows[yr] for yr in years_list if yr in year_rows]
                    if yr_refs:
                        cl = get_column_letter(ci)
                        refs_str = ','.join(f'{cl}{rr}' for rr in yr_refs)
                        if meth == 'AVERAGEIFS':
                            formula = f'=ROUND(AVERAGE({refs_str}),1)'
                        else:
                            formula = f'=ROUND(SUM({refs_str})/COUNT({refs_str}),1)'
                        c = ws.cell(row=r, column=ci, value=formula)
                        c.font = Font(name=FONT, size=9)
                        c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                        c.alignment = Alignment(horizontal='center', vertical='center')
                        c.border = _thin()
                        c.number_format = nf

        r += 2

        # ── 표 3: 이동평균 분석 ──────────────────
        if 'temp_avg' in self._col:
            tc = self._col['temp_avg']
            _title(ws, r, 1, 6, '표 3. 이동평균 분석 (연평균기온 기준)', h=20, bg=C['mid_blue'], sz=10)
            r += 1
            _hc(ws, r, 1, 'No.')
            _hc(ws, r, 2, '연도')
            _hc(ws, r, 3, '연평균기온(℃)')
            _hc(ws, r, 4, '5년 이동평균')
            _hc(ws, r, 5, '10년 이동평균')
            ws.column_dimensions['C'].width = 14
            ws.column_dimensions['D'].width = 14
            ws.column_dimensions['E'].width = 14

            ma_rows = {}
            for i, yr in enumerate(self._years):
                r += 1
                ma_rows[yr] = r
                bg = C['light_blue'] if i % 2 == 0 else C['white']
                _dc(ws, r, 1, i + 1, bg=bg)
                _dc(ws, r, 2, yr, bg=bg, bold=True)
                formula_yr = f'=ROUND(AVERAGEIFS({RAW_SHEET}!{tc}:{tc},{RAW_SHEET}!{yc}:{yc},{yr}),1)'
                c = ws.cell(row=r, column=3, value=formula_yr)
                c.font = Font(name=FONT, size=9)
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = _thin()
                c.number_format = '0.0'

                # 5년 이동평균
                if i >= 4:
                    prev5 = [ma_rows[self._years[j]] for j in range(i - 4, i + 1)]
                    refs5 = ','.join(f'C{rr}' for rr in prev5)
                    formula_5 = f'=ROUND(AVERAGE({refs5}),1)'
                else:
                    formula_5 = ''
                c5 = ws.cell(row=r, column=4, value=formula_5 if formula_5 else None)
                c5.font = Font(name=FONT, size=9)
                c5.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                c5.alignment = Alignment(horizontal='center', vertical='center')
                c5.border = _thin()
                c5.number_format = '0.0'

                # 10년 이동평균
                if i >= 9:
                    prev10 = [ma_rows[self._years[j]] for j in range(i - 9, i + 1)]
                    refs10 = ','.join(f'C{rr}' for rr in prev10)
                    formula_10 = f'=ROUND(AVERAGE({refs10}),1)'
                else:
                    formula_10 = ''
                c10 = ws.cell(row=r, column=5, value=formula_10 if formula_10 else None)
                c10.font = Font(name=FONT, size=9)
                c10.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                c10.alignment = Alignment(horizontal='center', vertical='center')
                c10.border = _thin()
                c10.number_format = '0.0'

            # 차트: 연평균기온 추이
            r += 2
            try:
                first_r = ma_rows[self._years[0]]
                last_r = ma_rows[self._years[-1]]
                lc = LineChart()
                lc.title = '연평균기온 추이 및 이동평균'
                lc.style = 10
                lc.y_axis.title = '기온(℃)'
                lc.x_axis.title = '연도'
                lc.width = 20
                lc.height = 12
                for col_n, col_title in [(3, '연평균기온'), (4, '5년이동평균'), (5, '10년이동평균')]:
                    data = Reference(ws, min_col=col_n, min_row=first_r - 1, max_row=last_r)
                    series = lc.series[len(lc.series)] if False else None
                    lc.add_data(data, titles_from_data=False)
                    lc.series[-1].title.value = col_title
                cats = Reference(ws, min_col=2, min_row=first_r, max_row=last_r)
                lc.set_categories(cats)
                ws.add_chart(lc, f'A{r}')
            except Exception:
                pass

    # ──────────────────────────────────
    # 5. 강수량 분석 시트
    # ──────────────────────────────────

    def _sheet_precipitation(self, wb, df):
        ws = wb.create_sheet("🌧️ 강수량 분석")
        ws.sheet_view.showGridLines = False

        if 'precipitation' not in self._col:
            _title(ws, 1, 1, 4, '강수량 분석 — 강수량 데이터 없음')
            return

        pc = self._col['precipitation']
        yc = self._col.get('year', 'B')
        mc = self._col.get('month', 'C')

        ws.column_dimensions['A'].width = 14

        r = 1
        _title(ws, r, 1, len(self._years) + 3, '강수량 분석')

        # ── 표 1: 연도별 기상개황 ──────────────────
        r = 2
        n_yr_cols = len(self._years)
        total_width = n_yr_cols + 3
        _title(ws, r, 1, total_width, '표 1. 연도별 기상개황', h=20, bg=C['mid_blue'], sz=10)
        r = 3
        _hc(ws, r, 1, '항목')
        for ci, yr in enumerate(self._years, 2):
            _hc(ws, r, ci, str(yr))
            ws.column_dimensions[get_column_letter(ci)].width = 10
        avg_col = n_yr_cols + 2
        _hc(ws, r, avg_col, '평균', bg=C['orange'], fg=C['white'])
        ws.column_dimensions[get_column_letter(avg_col)].width = 10

        temp_col = self._col.get('temp_avg')
        overview = [
            ('연평균기온(℃)', temp_col, 'AVERAGEIFS', '0.0'),
            ('총강수량(mm)',  pc, 'SUMIFS', '#,##0.0'),
            ('월평균강수량(mm)', pc, 'SUMIFS_12', '#,##0.0'),
        ]
        overview_rows = {}
        for item_name, ec, meth, nf in overview:
            r += 1
            _dc(ws, r, 1, item_name, bold=True, align='left')
            yr_cell_cols = []
            for ci, yr in enumerate(self._years, 2):
                if ec is None:
                    _dc(ws, r, ci, None)
                    continue
                if meth == 'AVERAGEIFS':
                    formula = f'=ROUND(AVERAGEIFS({RAW_SHEET}!{ec}:{ec},{RAW_SHEET}!{yc}:{yc},{yr}),1)'
                elif meth == 'SUMIFS':
                    formula = f'=ROUND(SUMIFS({RAW_SHEET}!{ec}:{ec},{RAW_SHEET}!{yc}:{yc},{yr}),1)'
                else:
                    formula = f'=ROUND(SUMIFS({RAW_SHEET}!{ec}:{ec},{RAW_SHEET}!{yc}:{yc},{yr})/12,1)'
                c = ws.cell(row=r, column=ci, value=formula)
                c.font = Font(name=FONT, size=9)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = _thin()
                c.number_format = nf
                yr_cell_cols.append(get_column_letter(ci))
            # 평균 열
            if yr_cell_cols:
                refs = ','.join(f'{cl}{r}' for cl in yr_cell_cols)
                avg_f = f'=ROUND(AVERAGE({refs}),1)'
                c = ws.cell(row=r, column=avg_col, value=avg_f)
                c.font = Font(name=FONT, size=9, bold=True)
                c.fill = PatternFill(start_color=C['light_orange'], end_color=C['light_orange'], fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = _thin()
                c.number_format = nf
            overview_rows[item_name] = r

        r += 2

        # ── 표 2: 연도별 월별 강수량 ──────────────────
        _title(ws, r, 1, total_width, '표 2. 연도별 월별 강수량 (mm)', h=20, bg=C['mid_blue'], sz=10)
        r += 1
        _hc(ws, r, 1, '월')
        for ci, yr in enumerate(self._years, 2):
            _hc(ws, r, ci, str(yr))
        _hc(ws, r, avg_col, '평균', bg=C['orange'], fg=C['white'])

        month_precip_rows = {}
        for m in range(1, 13):
            r += 1
            month_precip_rows[m] = r
            bg = C['light_blue'] if m % 2 == 0 else C['white']
            _dc(ws, r, 1, f'{m}월', bg=bg, bold=True)
            yr_cells = []
            for ci, yr in enumerate(self._years, 2):
                formula = f'=ROUND(SUMIFS({RAW_SHEET}!{pc}:{pc},{RAW_SHEET}!{yc}:{yc},{yr},{RAW_SHEET}!{mc}:{mc},{m}),1)'
                c = ws.cell(row=r, column=ci, value=formula)
                c.font = Font(name=FONT, size=9)
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = _thin()
                c.number_format = '#,##0.0'
                yr_cells.append(get_column_letter(ci))
            if yr_cells:
                refs = ','.join(f'{cl}{r}' for cl in yr_cells)
                c = ws.cell(row=r, column=avg_col, value=f'=ROUND(AVERAGE({refs}),1)')
                c.font = Font(name=FONT, size=9, bold=True)
                c.fill = PatternFill(start_color=C['light_orange'], end_color=C['light_orange'], fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = _thin()
                c.number_format = '#,##0.0'

        # 하단 집계
        first_m_r = month_precip_rows[1]
        last_m_r = month_precip_rows[12]
        for stat_name, bg, func in [
            ('평균', C['light_yellow'], 'AVERAGE'),
            ('최대', C['light_orange'], 'MAX'),
            ('최소', C['light_blue'], 'MIN'),
        ]:
            r += 1
            _dc(ws, r, 1, stat_name, bg=bg, bold=True)
            for ci in range(2, avg_col + 1):
                cl = get_column_letter(ci)
                formula = f'=ROUND({func}({cl}{first_m_r}:{cl}{last_m_r}),1)'
                c = ws.cell(row=r, column=ci, value=formula)
                c.font = Font(name=FONT, size=9, bold=True)
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = _thin()
                c.number_format = '#,##0.0'

        r += 2

        # ── 차트 ──────────────────
        try:
            # 연도별 연강수량 막대
            bc = BarChart()
            bc.title = '연도별 총 강수량'
            bc.style = 10
            bc.y_axis.title = '강수량(mm)'
            bc.x_axis.title = '연도'
            bc.width = 18
            bc.height = 10
            total_row = overview_rows.get('총강수량(mm)')
            if total_row:
                data = Reference(ws, min_col=2, max_col=n_yr_cols + 1, min_row=total_row)
                bc.add_data(data)
                cats = Reference(ws, min_col=2, max_col=n_yr_cols + 1, min_row=3)
                bc.set_categories(cats)
                ws.add_chart(bc, f'A{r}')

            # 월별 평균 강수량 막대
            bc2 = BarChart()
            bc2.title = '월별 평균 강수량'
            bc2.style = 10
            bc2.y_axis.title = '강수량(mm)'
            bc2.x_axis.title = '월'
            bc2.width = 18
            bc2.height = 10
            data2 = Reference(ws, min_col=avg_col, min_row=first_m_r - 1, max_row=last_m_r)
            bc2.add_data(data2, titles_from_data=True)
            cats2 = Reference(ws, min_col=1, min_row=first_m_r, max_row=last_m_r)
            bc2.set_categories(cats2)
            ws.add_chart(bc2, f'{get_column_letter(12)}{r}')
        except Exception:
            pass

    # ──────────────────────────────────
    # 6. 누적강수량 분석 시트
    # ──────────────────────────────────

    def _sheet_cumulative_precip(self, wb, df):
        ws = wb.create_sheet("누적강수량 분석")
        ws.freeze_panes = 'B2'

        if 'precipitation' not in self._col:
            _title(ws, 1, 1, 4, '누적강수량 분석 — 강수량 데이터 없음')
            return

        pc = self._col['precipitation']
        yc = self._col.get('year', 'B')
        mc = self._col.get('month', 'C')

        n_yr = len(self._years)
        # 이동평균 기준 연도 목록
        avg10_yrs = self._years[-10:] if len(self._years) >= 10 else self._years
        avg20_yrs = self._years[-20:] if len(self._years) >= 20 else self._years
        avg30_yrs = self._years[-30:] if len(self._years) >= 30 else self._years

        # 컬럼 설정: 항목, 연도들, 10년평균, 20년평균, 30년평균
        ws.column_dimensions['A'].width = 14
        for ci in range(2, n_yr + 5):
            ws.column_dimensions[get_column_letter(ci)].width = 9

        yr_start_col = 2
        avg10_col = n_yr + 2
        avg20_col = n_yr + 3
        avg30_col = n_yr + 4

        def write_header_row(r, section_title):
            _hc(ws, r, 1, section_title, bg=C['mid_blue'])
            for ci, yr in enumerate(self._years, yr_start_col):
                _hc(ws, r, ci, str(yr), sz=8)
            _hc(ws, r, avg10_col, f'10년평균\n({avg10_yrs[0]}~)', sz=8, wrap=True, bg=C['orange'])
            _hc(ws, r, avg20_col, f'20년평균\n({avg20_yrs[0]}~)', sz=8, wrap=True, bg=C['orange'])
            _hc(ws, r, avg30_col, f'30년평균\n({avg30_yrs[0]}~)', sz=8, wrap=True, bg=C['orange'])
            ws.row_dimensions[r].height = 28

        # 여름 조합 정의
        summer_combos = [
            ('6+7월',   [6, 7]),
            ('6+7+8월', [6, 7, 8]),
            ('7+8월',   [7, 8]),
            ('7+8+9월', [7, 8, 9]),
        ]

        # ── 구간①: 연도별 월별 강수량 (절대값) ──────────────────
        r = 1
        write_header_row(r, '월별 강수량(mm)')

        monthly_cell = {}  # (yr_idx, month) → cell address

        for m in range(1, 13):
            r += 1
            bg = C['light_blue'] if m % 2 == 0 else C['white']
            _dc(ws, r, 1, f'{m}월', bg=bg, bold=True)
            yr_cells = []
            for ci, yr in enumerate(self._years, yr_start_col):
                formula = f'=ROUND(SUMIFS({RAW_SHEET}!{pc}:{pc},{RAW_SHEET}!{yc}:{yc},{yr},{RAW_SHEET}!{mc}:{mc},{m}),1)'
                c = ws.cell(row=r, column=ci, value=formula)
                c.font = Font(name=FONT, size=9)
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = _thin()
                c.number_format = '#,##0.0'
                monthly_cell[(ci - yr_start_col, m)] = (get_column_letter(ci), r)
                yr_cells.append((get_column_letter(ci), r))

            # 이동평균 열
            def _avg_formula(yrs_list):
                idxs = [i for i, yr in enumerate(self._years) if yr in yrs_list]
                if not idxs:
                    return ''
                refs = ','.join(f'{get_column_letter(idx + yr_start_col)}{r}' for idx in idxs)
                return f'=ROUND(AVERAGE({refs}),1)'

            for avg_col, avg_yrs in [(avg10_col, avg10_yrs), (avg20_col, avg20_yrs), (avg30_col, avg30_yrs)]:
                formula = _avg_formula(avg_yrs)
                c = ws.cell(row=r, column=avg_col, value=formula if formula else None)
                c.font = Font(name=FONT, size=9)
                c.fill = PatternFill(start_color=C['light_orange'], end_color=C['light_orange'], fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = _thin()
                c.number_format = '#,##0.0'

        # 합계 행
        month_rows_sec1 = {m: r - (12 - m) for m in range(1, 13)}
        # 재계산: 각 월 row 기억
        month_rows_abs = {}
        row_ptr = 2
        for m in range(1, 13):
            month_rows_abs[m] = row_ptr
            row_ptr += 1

        r += 1
        _dc(ws, r, 1, '연합계', bg=C['yellow'], bold=True)
        for ci in range(yr_start_col, avg30_col + 1):
            cl = get_column_letter(ci)
            refs = ','.join(f'{cl}{month_rows_abs[m]}' for m in range(1, 13))
            formula = f'=ROUND(SUM({refs}),1)'
            c = ws.cell(row=r, column=ci, value=formula)
            c.font = Font(name=FONT, size=9, bold=True)
            c.fill = PatternFill(start_color=C['yellow'], end_color=C['yellow'], fill_type='solid')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = _thin()
            c.number_format = '#,##0.0'
        annual_abs_row = r

        # 여름 조합 행
        for combo_name, months in summer_combos:
            r += 1
            _dc(ws, r, 1, combo_name, bg=C['light_green'], bold=True)
            for ci in range(yr_start_col, avg30_col + 1):
                cl = get_column_letter(ci)
                refs = ','.join(f'{cl}{month_rows_abs[m]}' for m in months)
                formula = f'=ROUND(SUM({refs}),1)'
                c = ws.cell(row=r, column=ci, value=formula)
                c.font = Font(name=FONT, size=9)
                c.fill = PatternFill(start_color=C['light_green'], end_color=C['light_green'], fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = _thin()
                c.number_format = '#,##0.0'

        r += 2

        # ── 구간②: 월별 누적강수량 ──────────────────
        cum_start_row = r
        write_header_row(r, '누적강수량(mm)')

        cum_rows = {}
        for m in range(1, 13):
            r += 1
            cum_rows[m] = r
            bg = C['light_blue'] if m % 2 == 0 else C['white']
            _dc(ws, r, 1, f'1~{m}월', bg=bg, bold=True)
            for ci in range(yr_start_col, avg30_col + 1):
                cl = get_column_letter(ci)
                refs = ','.join(f'{cl}{month_rows_abs[mm]}' for mm in range(1, m + 1))
                formula = f'=ROUND(SUM({refs}),1)'
                c = ws.cell(row=r, column=ci, value=formula)
                c.font = Font(name=FONT, size=9)
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = _thin()
                c.number_format = '#,##0.0'

        # 여름 조합 누적
        for combo_name, months in summer_combos:
            r += 1
            _dc(ws, r, 1, f'누적 {combo_name}', bg=C['light_green'], bold=True)
            for ci in range(yr_start_col, avg30_col + 1):
                cl = get_column_letter(ci)
                refs = ','.join(f'{cl}{month_rows_abs[m]}' for m in months)
                formula = f'=ROUND(SUM({refs}),1)'
                c = ws.cell(row=r, column=ci, value=formula)
                c.font = Font(name=FONT, size=9)
                c.fill = PatternFill(start_color=C['light_green'], end_color=C['light_green'], fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = _thin()
                c.number_format = '#,##0.0'

        r += 2

        # ── 구간③: 연간 대비 누적강수량 비율 ──────────────────
        write_header_row(r, '누적비율(%)')
        ratio_start = r

        for m in range(1, 13):
            r += 1
            bg = C['light_blue'] if m % 2 == 0 else C['white']
            _dc(ws, r, 1, f'1~{m}월 비율', bg=bg, bold=True)
            for ci in range(yr_start_col, avg30_col + 1):
                cl = get_column_letter(ci)
                cum_r = cum_rows[m]
                ann_r = annual_abs_row
                formula = f'=IFERROR(ROUND({cl}{cum_r}/{cl}{ann_r}*100,1),"")'
                c = ws.cell(row=r, column=ci, value=formula)
                c.font = Font(name=FONT, size=9)
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = _thin()
                c.number_format = '0.0'

    # ──────────────────────────────────
    # 7. 강우일수 분석 시트
    # ──────────────────────────────────

    def _sheet_rainfall_days(self, wb, df):
        ws = wb.create_sheet("강우일수 분석")
        ws.freeze_panes = 'A2'

        if 'precipitation' not in df.columns:
            _title(ws, 1, 1, 4, '강우일수 분석 — 강수량 데이터 없음')
            return

        col_headers = [
            '연도', '강우일수\n(>0mm)', '무강우일수\n(=0mm)', '총일수',
            '< 3mm', '3~10mm', '10~30mm', '30~80mm', '≥ 80mm',
            '최장연속\n강우(일)', '최장연속\n무강우(일)',
        ]
        n_cols = len(col_headers)

        stations = df['station_name'].unique() if self._has_stn else ['전체']
        r = 1

        for stn in stations:
            sdf = df[df['station_name'] == stn].copy() if self._has_stn else df.copy()

            if 'year' not in sdf.columns:
                continue

            # 섹션 타이틀
            title_text = f'강우일수 분석 — {stn}' if self._has_stn else '강우일수 분석'
            _title(ws, r, 1, n_cols, title_text, h=22, bg=C['dark_blue'], sz=11)
            r += 1

            # 헤더
            for ci, h in enumerate(col_headers, 1):
                _hc(ws, r, ci, h, sz=9, wrap=True)
                ws.column_dimensions[get_column_letter(ci)].width = 10
            ws.row_dimensions[r].height = 28
            r += 1

            records = []
            for yr in sorted(sdf['year'].dropna().unique().astype(int)):
                ydf = sdf[sdf['year'] == yr].copy()
                prec = ydf['precipitation'].fillna(0)

                rain_days = int((prec > 0).sum())
                no_rain_days = int((prec == 0).sum())
                total_days = len(prec)

                lt3 = int(((prec > 0) & (prec < 3)).sum())
                r3_10 = int(((prec >= 3) & (prec < 10)).sum())
                r10_30 = int(((prec >= 10) & (prec < 30)).sum())
                r30_80 = int(((prec >= 30) & (prec < 80)).sum())
                ge80 = int((prec >= 80).sum())

                # 최장 연속 강우
                max_consec_rain = 0
                cur = 0
                for v in prec:
                    if v > 0:
                        cur += 1
                        max_consec_rain = max(max_consec_rain, cur)
                    else:
                        cur = 0

                # 최장 연속 무강우
                max_consec_dry = 0
                cur = 0
                for v in prec:
                    if v == 0:
                        cur += 1
                        max_consec_dry = max(max_consec_dry, cur)
                    else:
                        cur = 0

                records.append([
                    yr, rain_days, no_rain_days, total_days,
                    lt3, r3_10, r10_30, r30_80, ge80,
                    max_consec_rain, max_consec_dry,
                ])

            data_start_row = r
            for i, rec in enumerate(records):
                bg = C['light_blue'] if i % 2 == 0 else C['white']
                for ci, val in enumerate(rec, 1):
                    _dc(ws, r, ci, val, bg=bg)
                r += 1
            data_end_row = r - 1

            # 집계 행
            if records:
                for stat_name, bg, func in [
                    ('평균', C['light_yellow'], 'AVERAGE'),
                    ('최대', C['light_orange'], 'MAX'),
                    ('최소', C['light_blue'], 'MIN'),
                ]:
                    _dc(ws, r, 1, stat_name, bg=bg, bold=True)
                    for ci in range(2, n_cols + 1):
                        cl = get_column_letter(ci)
                        formula = f'=ROUND({func}({cl}{data_start_row}:{cl}{data_end_row}),1)'
                        c = ws.cell(row=r, column=ci, value=formula)
                        c.font = Font(name=FONT, size=9, bold=True)
                        c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                        c.alignment = Alignment(horizontal='center', vertical='center')
                        c.border = _thin()
                        c.number_format = '0.0'
                    r += 1

            r += 2

    # ──────────────────────────────────
    # 8. 피벗 분석 시트
    # ──────────────────────────────────

    def _sheet_pivot(self, wb, df):
        ws = wb.create_sheet("🔄 피벗 분석")
        ws.sheet_view.showGridLines = True
        ws.freeze_panes = 'C3'

        yc = self._col.get('year', 'B')
        mc = self._col.get('month', 'C')

        pivot_elems = [(k, lbl) for k, lbl in ELEMENT_LABELS.items() if k in self._col]
        if not pivot_elems:
            _title(ws, 1, 1, 4, '피벗 분석 — 데이터 없음')
            return

        # 컬럼 폭
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 8
        for ci in range(3, 16):
            ws.column_dimensions[get_column_letter(ci)].width = 9

        r = 1
        for key, elem_lbl in pivot_elems:
            ec = self._col[key]
            is_sum = key in SUM_ELEMENTS

            # 요소 타이틀
            _title(ws, r, 1, 15, f'▶ {elem_lbl}  ({"월합산" if is_sum else "월평균"})', h=20, bg=C['dark_blue'], sz=10)
            r += 1

            # 헤더: 연도, 구분, 1~12월, 연간
            _hc(ws, r, 1, '연도', sz=9)
            _hc(ws, r, 2, '구분', sz=9)
            for m in range(1, 13):
                _hc(ws, r, m + 2, f'{m}월', sz=9)
            _hc(ws, r, 15, '연간', bg=C['orange'], fg=C['white'], sz=9)
            r += 1

            year_block_rows = {}
            for yr in self._years:
                year_block_rows[yr] = r
                bg = C['light_blue']
                _dc(ws, r, 1, yr, bg=bg, bold=True)
                _dc(ws, r, 2, '값', bg=bg)
                month_cells = []
                for m in range(1, 13):
                    if is_sum:
                        formula = (f'=IFERROR(ROUND(SUMIFS({RAW_SHEET}!{ec}:{ec},'
                                   f'{RAW_SHEET}!{yc}:{yc},{yr},'
                                   f'{RAW_SHEET}!{mc}:{mc},{m}),1),"")')
                    else:
                        formula = (f'=IFERROR(ROUND(AVERAGEIFS({RAW_SHEET}!{ec}:{ec},'
                                   f'{RAW_SHEET}!{yc}:{yc},{yr},'
                                   f'{RAW_SHEET}!{mc}:{mc},{m}),1),"")')
                    c = ws.cell(row=r, column=m + 2, value=formula)
                    c.font = Font(name=FONT, size=9)
                    c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border = _thin()
                    c.number_format = '#,##0.0'
                    month_cells.append(f'{get_column_letter(m + 2)}{r}')

                # 연간 열
                refs = ','.join(month_cells)
                if is_sum:
                    annual_f = f'=IFERROR(ROUND(SUM({refs}),1),"")'
                else:
                    annual_f = f'=IFERROR(ROUND(AVERAGE({refs}),1),"")'
                c = ws.cell(row=r, column=15, value=annual_f)
                c.font = Font(name=FONT, size=9, bold=True)
                c.fill = PatternFill(start_color=C['light_orange'], end_color=C['light_orange'], fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = _thin()
                c.number_format = '#,##0.0'
                r += 1

            # 집계 행 (전체평균, 최대, 최소)
            if year_block_rows:
                first_yr_r = year_block_rows[self._years[0]]
                last_yr_r = year_block_rows[self._years[-1]]
                for stat_name, bg, func in [
                    ('전체평균', C['light_yellow'], 'AVERAGE'),
                    ('최대', C['light_orange'], 'MAX'),
                    ('최소', C['light_blue'], 'MIN'),
                ]:
                    _dc(ws, r, 1, stat_name, bg=bg, bold=True)
                    _dc(ws, r, 2, func, bg=bg)
                    for ci in range(3, 16):
                        cl = get_column_letter(ci)
                        formula = f'=IFERROR(ROUND({func}({cl}{first_yr_r}:{cl}{last_yr_r}),1),"")'
                        c = ws.cell(row=r, column=ci, value=formula)
                        c.font = Font(name=FONT, size=9, bold=True)
                        c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                        c.alignment = Alignment(horizontal='center', vertical='center')
                        c.border = _thin()
                        c.number_format = '#,##0.0'
                    r += 1

            r += 2

    # ──────────────────────────────────
    # 9. 기상개황 시트 (표 2-1)
    # ──────────────────────────────────

    def _sheet_weather_overview(self, wb, df):
        ws = wb.create_sheet("📋 기상개황")
        ws.sheet_view.showGridLines = False

        stations = df['station_name'].unique() if self._has_stn else ['전체']
        years = sorted(df['year'].dropna().unique().astype(int).tolist())
        n_years = len(years)

        # 컬럼 폭 설정: A=항목(18), B=구분(8), C~(연도별 8), 마지막=평균(10)
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 8
        for ci in range(3, 3 + n_years):
            ws.column_dimensions[get_column_letter(ci)].width = 8
        avg_col = 3 + n_years
        ws.column_dimensions[get_column_letter(avg_col)].width = 10

        yr_min = years[0] if years else '-'
        yr_max = years[-1] if years else '-'

        r = 1
        total_width = 2 + n_years + 1
        _title(ws, r, 1, total_width, '기상개황', h=28, sz=14)
        r += 1

        for stn in stations:
            sdf = df[df['station_name'] == stn].copy() if self._has_stn else df.copy()

            if self._has_stn:
                _title(ws, r, 1, total_width,
                       f'▶ {stn} 관측소', h=22, bg=C['mid_blue'], sz=11)
                r += 1

            _title(ws, r, 1, total_width,
                   '표 2-1. 연도별 기상개황', h=20, bg=C['mid_blue'], sz=10)
            r += 1

            # 헤더
            _hc(ws, r, 1, '항목', bg=C['dark_blue'])
            _hc(ws, r, 2, '구분', bg=C['dark_blue'])
            for ci, yr in enumerate(years, 3):
                _hc(ws, r, ci, str(yr), bg=C['dark_blue'])
            _hc(ws, r, avg_col, '평균', bg=C['dark_blue'])
            r += 1

            # ── 기온 블록 ──
            temp_items = [
                ('temp_avg', '평균'),
                ('temp_max', '최고'),
                ('temp_min', '최저'),
            ]
            avail_temp = [(k, sl) for k, sl in temp_items if k in sdf.columns]

            if avail_temp:
                bg = C['light_blue']
                merge_end = r + len(avail_temp) - 1
                ws.merge_cells(start_row=r, start_column=1, end_row=merge_end, end_column=1)
                c = ws.cell(row=r, column=1, value='평균기온(℃)')
                c.font = Font(name=FONT, bold=True, size=9, color='000000')
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                c.border = _thin()
                for merge_r in range(r + 1, merge_end + 1):
                    ws.cell(row=merge_r, column=1).border = _thin()
                    ws.cell(row=merge_r, column=1).fill = PatternFill(
                        start_color=bg, end_color=bg, fill_type='solid')

                for col_key, sub_label in avail_temp:
                    _dc(ws, r, 2, sub_label, bg=bg, bold=True)
                    yr_vals = []
                    for ci, yr in enumerate(years, 3):
                        ydf = sdf[sdf['year'] == yr]
                        v = _safe_val(ydf[col_key].mean() if not ydf.empty else None)
                        val = round(float(v), 1) if v is not None else None
                        _dc(ws, r, ci, val, bg=bg, nf='0.0')
                        yr_vals.append(val)
                    valid_vals = [v for v in yr_vals if v is not None]
                    avg_val = round(sum(valid_vals) / len(valid_vals), 1) if valid_vals else None
                    _dc(ws, r, avg_col, avg_val, bg=bg, bold=True, nf='0.0')
                    r += 1

            # ── 평균월강수량 ──
            if 'precipitation' in sdf.columns:
                bg = C['light_green']
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
                c = ws.cell(row=r, column=1, value='평균월강수량(mm)')
                c.font = Font(name=FONT, bold=True, size=9, color='000000')
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = _thin()
                ws.cell(row=r, column=2).border = _thin()
                ws.cell(row=r, column=2).fill = PatternFill(
                    start_color=bg, end_color=bg, fill_type='solid')
                yr_vals = []
                for ci, yr in enumerate(years, 3):
                    ydf = sdf[sdf['year'] == yr]
                    v = _safe_val(ydf['precipitation'].sum() if not ydf.empty else None)
                    val = round(float(v) / 12, 1) if v is not None else None
                    _dc(ws, r, ci, val, bg=bg, nf='#,##0.0')
                    yr_vals.append(val)
                valid_vals = [v for v in yr_vals if v is not None]
                avg_val = round(sum(valid_vals) / len(valid_vals), 1) if valid_vals else None
                _dc(ws, r, avg_col, avg_val, bg=bg, bold=True, nf='#,##0.0')
                r += 1

            # ── 평균풍속 ──
            if 'wind_speed' in sdf.columns:
                bg = C['light_gray']
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
                c = ws.cell(row=r, column=1, value='평균풍속(m/s)')
                c.font = Font(name=FONT, bold=True, size=9, color='000000')
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = _thin()
                ws.cell(row=r, column=2).border = _thin()
                ws.cell(row=r, column=2).fill = PatternFill(
                    start_color=bg, end_color=bg, fill_type='solid')
                yr_vals = []
                for ci, yr in enumerate(years, 3):
                    ydf = sdf[sdf['year'] == yr]
                    v = _safe_val(ydf['wind_speed'].mean() if not ydf.empty else None)
                    val = round(float(v), 1) if v is not None else None
                    _dc(ws, r, ci, val, bg=bg, nf='0.0')
                    yr_vals.append(val)
                valid_vals = [v for v in yr_vals if v is not None]
                avg_val = round(sum(valid_vals) / len(valid_vals), 1) if valid_vals else None
                _dc(ws, r, avg_col, avg_val, bg=bg, bold=True, nf='0.0')
                r += 1

            # ── 최대순간풍속 ──
            gust_col = ('wind_gust' if 'wind_gust' in sdf.columns
                        else ('wind_max' if 'wind_max' in sdf.columns else None))
            if gust_col:
                bg = C['light_gray']
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
                c = ws.cell(row=r, column=1, value='최대순간풍속(m/s)')
                c.font = Font(name=FONT, bold=True, size=9, color='000000')
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = _thin()
                ws.cell(row=r, column=2).border = _thin()
                ws.cell(row=r, column=2).fill = PatternFill(
                    start_color=bg, end_color=bg, fill_type='solid')
                yr_vals = []
                for ci, yr in enumerate(years, 3):
                    ydf = sdf[sdf['year'] == yr]
                    v = _safe_val(ydf[gust_col].max() if not ydf.empty else None)
                    val = round(float(v), 1) if v is not None else None
                    _dc(ws, r, ci, val, bg=bg, nf='0.0')
                    yr_vals.append(val)
                valid_vals = [v for v in yr_vals if v is not None]
                avg_val = round(sum(valid_vals) / len(valid_vals), 1) if valid_vals else None
                _dc(ws, r, avg_col, avg_val, bg=bg, bold=True, nf='0.0')
                r += 1

            r += 1

            # ── 강수현황 구분 헤더 ──
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_width)
            c = ws.cell(row=r, column=1, value='강수현황(mm)')
            c.font = Font(name=FONT, bold=True, size=10, color=C['white'])
            c.fill = PatternFill(start_color=C['mid_blue'], end_color=C['mid_blue'], fill_type='solid')
            c.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[r].height = 20
            r += 1

            # ── 강수 피벗 헤더 ──
            _hc(ws, r, 1, '년', bg=C['dark_blue'])
            for mi in range(1, 13):
                _hc(ws, r, mi + 1, f'{mi}월', bg=C['dark_blue'])
            sum_col = 14
            _hc(ws, r, sum_col, '합계', bg=C['dark_blue'])
            r += 1

            if 'precipitation' in sdf.columns:
                # 피벗 데이터 계산
                pivot_data = {}
                for yr in years:
                    ydf = sdf[sdf['year'] == yr]
                    monthly_sums = {}
                    for m in range(1, 13):
                        mdf = ydf[ydf['month'] == m]
                        v = _safe_val(mdf['precipitation'].sum() if not mdf.empty else 0.0)
                        monthly_sums[m] = round(float(v), 1) if v is not None else 0.0
                    pivot_data[yr] = monthly_sums

                for yi, yr in enumerate(years):
                    bg = C['light_green'] if yi % 2 == 0 else C['white']
                    _dc(ws, r, 1, yr, bg=bg, bold=True)
                    row_vals = []
                    for mi, m in enumerate(range(1, 13), 2):
                        val = pivot_data[yr].get(m, 0.0)
                        _dc(ws, r, mi, val, bg=bg, nf='#,##0.0')
                        row_vals.append(val)
                    total = round(sum(v for v in row_vals if v is not None), 1)
                    _dc(ws, r, sum_col, total, bg=bg, bold=True, nf='#,##0.0')
                    r += 1

                # 통계 행
                for stat_name, bg, agg_fn in [
                    ('평균', C['light_yellow'],
                     lambda vals: round(sum(vals) / len(vals), 1) if vals else None),
                    ('최대', C['light_orange'],
                     lambda vals: round(max(vals), 1) if vals else None),
                    ('최소', C['light_blue'],
                     lambda vals: round(min(vals), 1) if vals else None),
                ]:
                    _dc(ws, r, 1, stat_name, bg=bg, bold=True)
                    row_totals = []
                    for mi, m in enumerate(range(1, 13), 2):
                        col_vals_valid = [
                            pivot_data[yr].get(m, 0.0)
                            for yr in years
                            if pivot_data[yr].get(m) is not None
                        ]
                        val = agg_fn(col_vals_valid)
                        _dc(ws, r, mi, val, bg=bg, bold=True, nf='#,##0.0')
                        if val is not None:
                            row_totals.append(val)
                    total_val = round(sum(row_totals), 1) if row_totals else None
                    _dc(ws, r, sum_col, total_val, bg=bg, bold=True, nf='#,##0.0')
                    r += 1

            r += 1
            _note(ws, r, 1, total_width,
                  f'※ 자료 : 기상자료개방포털({yr_min}~{yr_max})')
            r += 1

            # ── 내장 차트 ──
            chart_anchor_row = r + 2

            # 차트1: 연도별 기온 꺾은선 (임시 데이터 영역 활용)
            try:
                if any(k in sdf.columns for k in ('temp_avg', 'temp_max', 'temp_min')):
                    temp_chart_col = total_width + 2
                    ws.cell(row=chart_anchor_row, column=temp_chart_col, value='연도')
                    ws.cell(row=chart_anchor_row, column=temp_chart_col + 1, value='평균기온')
                    ws.cell(row=chart_anchor_row, column=temp_chart_col + 2, value='최고기온')
                    ws.cell(row=chart_anchor_row, column=temp_chart_col + 3, value='최저기온')

                    for yi, yr in enumerate(years):
                        row_idx = chart_anchor_row + 1 + yi
                        ws.cell(row=row_idx, column=temp_chart_col, value=yr)
                        for col_offset, col_key in enumerate(
                                ['temp_avg', 'temp_max', 'temp_min'], 1):
                            if col_key in sdf.columns:
                                ydf = sdf[sdf['year'] == yr]
                                v = _safe_val(ydf[col_key].mean() if not ydf.empty else None)
                                ws.cell(row=row_idx, column=temp_chart_col + col_offset,
                                        value=round(float(v), 1) if v is not None else None)

                    data_end_row = chart_anchor_row + len(years)
                    lc = LineChart()
                    lc.title = '연도별 기온 추이'
                    lc.style = 10
                    lc.y_axis.title = '기온(℃)'
                    lc.x_axis.title = '연도'
                    lc.width = 20
                    lc.height = 12
                    for col_offset in range(1, 4):
                        data = Reference(ws,
                                         min_col=temp_chart_col + col_offset,
                                         min_row=chart_anchor_row,
                                         max_row=data_end_row)
                        lc.add_data(data, titles_from_data=True)
                    cats = Reference(ws,
                                     min_col=temp_chart_col,
                                     min_row=chart_anchor_row + 1,
                                     max_row=data_end_row)
                    lc.set_categories(cats)
                    ws.add_chart(lc, f'D{chart_anchor_row}')
            except Exception:
                pass

            # 차트2: 연강수량 막대 + 월평균강수량 꺾은선 콤보
            try:
                if 'precipitation' in sdf.columns:
                    precip_chart_col = total_width + 7
                    ws.cell(row=chart_anchor_row, column=precip_chart_col, value='연도')
                    ws.cell(row=chart_anchor_row, column=precip_chart_col + 1, value='연강수량')
                    ws.cell(row=chart_anchor_row, column=precip_chart_col + 2, value='월평균강수량')

                    for yi, yr in enumerate(years):
                        row_idx = chart_anchor_row + 1 + yi
                        ydf = sdf[sdf['year'] == yr]
                        ws.cell(row=row_idx, column=precip_chart_col, value=yr)
                        total_p = _safe_val(ydf['precipitation'].sum() if not ydf.empty else None)
                        total_p_val = round(float(total_p), 1) if total_p is not None else None
                        monthly_p = round(float(total_p) / 12, 1) if total_p is not None else None
                        ws.cell(row=row_idx, column=precip_chart_col + 1, value=total_p_val)
                        ws.cell(row=row_idx, column=precip_chart_col + 2, value=monthly_p)

                    data2_end_row = chart_anchor_row + len(years)

                    bc2 = BarChart()
                    bc2.title = '연도별 강수량'
                    bc2.style = 10
                    bc2.y_axis.title = '연강수량(mm)'
                    bc2.x_axis.title = '연도'
                    bc2.width = 20
                    bc2.height = 12
                    bar_data = Reference(ws,
                                         min_col=precip_chart_col + 1,
                                         min_row=chart_anchor_row,
                                         max_row=data2_end_row)
                    bc2.add_data(bar_data, titles_from_data=True)

                    lc2 = LineChart()
                    lc2.y_axis.axId = 200
                    lc2.y_axis.title = '월평균강수량(mm)'
                    line_data = Reference(ws,
                                          min_col=precip_chart_col + 2,
                                          min_row=chart_anchor_row,
                                          max_row=data2_end_row)
                    lc2.add_data(line_data, titles_from_data=True)
                    bc2 += lc2

                    cats2 = Reference(ws,
                                      min_col=precip_chart_col,
                                      min_row=chart_anchor_row + 1,
                                      max_row=data2_end_row)
                    bc2.set_categories(cats2)
                    ws.add_chart(bc2, f'D{chart_anchor_row + 22}')
            except Exception:
                pass

            r = chart_anchor_row + 50  # 다음 관측소 블록을 위한 여백

    # ──────────────────────────────────
    # 10. 기후통계 시트 (표 2-2)
    # ──────────────────────────────────

    def _sheet_climate_monthly(self, wb, df):
        ws = wb.create_sheet("📊 기후통계")
        ws.sheet_view.showGridLines = False

        stations = df['station_name'].unique() if self._has_stn else ['전체']
        years = sorted(df['year'].dropna().unique().astype(int).tolist())

        yr_min = years[0] if years else '-'
        yr_max = years[-1] if years else '-'

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        for ci in range(3, 15):
            ws.column_dimensions[get_column_letter(ci)].width = 8
        ws.column_dimensions['O'].width = 12

        total_width = 15

        r = 1
        _title(ws, r, 1, total_width, '기후통계', h=28, sz=14)
        r += 1

        for stn in stations:
            sdf = df[df['station_name'] == stn].copy() if self._has_stn else df.copy()

            if self._has_stn:
                _title(ws, r, 1, total_width,
                       f'▶ {stn} 관측소', h=22, bg=C['mid_blue'], sz=11)
                r += 1

            _title(ws, r, 1, total_width,
                   '표 2-2. 월별 기후통계', h=20, bg=C['mid_blue'], sz=10)
            r += 1

            _hc(ws, r, 1, '구분', bg=C['dark_blue'])
            _hc(ws, r, 2, '소구분', bg=C['dark_blue'])
            for mi, m in enumerate(range(1, 13), 3):
                _hc(ws, r, mi, f'{m}월', bg=C['dark_blue'])
            _hc(ws, r, 15, '평균/합계', bg=C['dark_blue'])
            r += 1

            def _monthly_vals(col_key, agg='mean'):
                result = {}
                if col_key not in sdf.columns:
                    return result
                for m in range(1, 13):
                    mdf = sdf[sdf['month'] == m]
                    if mdf.empty:
                        result[m] = None
                        continue
                    if agg == 'mean':
                        v = mdf[col_key].mean()
                    elif agg == 'max':
                        v = mdf[col_key].max()
                    elif agg == 'min':
                        v = mdf[col_key].min()
                    elif agg == 'sum':
                        v = mdf[col_key].sum()
                    else:
                        v = None
                    v = _safe_val(v)
                    result[m] = round(float(v), 1) if v is not None else None
                return result

            # 기온 블록
            temp_defs = [
                ('temp_avg', '평균',    'mean', '평균'),
                ('temp_max', '평균최고', 'mean', '평균'),
                ('temp_max', '최고극값', 'max',  '최대'),
                ('temp_min', '평균최저', 'mean', '평균'),
                ('temp_min', '최저극값', 'min',  '최소'),
            ]
            avail_temp = [(ck, sl, ag, ft) for ck, sl, ag, ft in temp_defs
                          if ck in sdf.columns]

            if avail_temp:
                bg = C['light_blue']
                merge_end = r + len(avail_temp) - 1
                ws.merge_cells(start_row=r, start_column=1, end_row=merge_end, end_column=1)
                c = ws.cell(row=r, column=1, value='기온(℃)')
                c.font = Font(name=FONT, bold=True, size=9, color='000000')
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                c.border = _thin()
                for merge_r in range(r + 1, merge_end + 1):
                    ws.cell(row=merge_r, column=1).border = _thin()
                    ws.cell(row=merge_r, column=1).fill = PatternFill(
                        start_color=bg, end_color=bg, fill_type='solid')

                for col_key, sub_label, agg, final_type in avail_temp:
                    monthly = _monthly_vals(col_key, agg)
                    valid = [v for v in monthly.values() if v is not None]
                    if final_type == '평균':
                        last_val = round(sum(valid) / len(valid), 1) if valid else None
                    elif final_type == '최대':
                        last_val = round(max(valid), 1) if valid else None
                    else:
                        last_val = round(min(valid), 1) if valid else None
                    _dc(ws, r, 2, sub_label, bg=bg)
                    for mi, m in enumerate(range(1, 13), 3):
                        _dc(ws, r, mi, monthly.get(m), bg=bg, nf='0.0')
                    _dc(ws, r, 15, last_val, bg=bg, bold=True, nf='0.0')
                    r += 1

            # 강수량 블록
            precip_defs = [
                ('precipitation', '평균', 'mean', '#,##0.0'),
                ('precipitation', '최대', 'max',  '#,##0.0'),
                ('precipitation', '최소', 'min',  '#,##0.0'),
            ]
            avail_precip = [(ck, sl, ag, nf) for ck, sl, ag, nf in precip_defs
                            if ck in sdf.columns]

            if avail_precip:
                bg = C['light_green']
                merge_end = r + len(avail_precip) - 1
                ws.merge_cells(start_row=r, start_column=1, end_row=merge_end, end_column=1)
                c = ws.cell(row=r, column=1, value='강수량(mm)')
                c.font = Font(name=FONT, bold=True, size=9, color='000000')
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                c.border = _thin()
                for merge_r in range(r + 1, merge_end + 1):
                    ws.cell(row=merge_r, column=1).border = _thin()
                    ws.cell(row=merge_r, column=1).fill = PatternFill(
                        start_color=bg, end_color=bg, fill_type='solid')

                for col_key, sub_label, agg, nf in avail_precip:
                    monthly = _monthly_vals(col_key, agg)
                    valid = [v for v in monthly.values() if v is not None]
                    if sub_label == '평균':
                        yr_totals = []
                        for yr in years:
                            ydf = sdf[sdf['year'] == yr]
                            v = _safe_val(ydf['precipitation'].sum() if not ydf.empty else None)
                            if v is not None:
                                yr_totals.append(float(v))
                        last_val = round(sum(yr_totals) / len(yr_totals), 1) if yr_totals else None
                    elif sub_label == '최대':
                        last_val = round(max(valid), 1) if valid else None
                    else:
                        last_val = round(min(valid), 1) if valid else None
                    _dc(ws, r, 2, sub_label, bg=bg)
                    for mi, m in enumerate(range(1, 13), 3):
                        _dc(ws, r, mi, monthly.get(m), bg=bg, nf=nf)
                    _dc(ws, r, 15, last_val, bg=bg, bold=True, nf=nf)
                    r += 1

            # 상대습도 블록
            if 'humidity_min' in sdf.columns:
                hum_defs = [('humidity', '평균', 'mean'), ('humidity_min', '최소', 'min')]
            elif 'humidity' in sdf.columns:
                hum_defs = [('humidity', '평균', 'mean'), ('humidity', '최소', 'min')]
            else:
                hum_defs = []
            avail_hum = [(ck, sl, ag) for ck, sl, ag in hum_defs if ck in sdf.columns]

            if avail_hum:
                bg = C['light_yellow']
                merge_end = r + len(avail_hum) - 1
                ws.merge_cells(start_row=r, start_column=1, end_row=merge_end, end_column=1)
                c = ws.cell(row=r, column=1, value='상대습도(%)')
                c.font = Font(name=FONT, bold=True, size=9, color='000000')
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                c.border = _thin()
                for merge_r in range(r + 1, merge_end + 1):
                    ws.cell(row=merge_r, column=1).border = _thin()
                    ws.cell(row=merge_r, column=1).fill = PatternFill(
                        start_color=bg, end_color=bg, fill_type='solid')

                for col_key, sub_label, agg in avail_hum:
                    monthly = _monthly_vals(col_key, agg)
                    valid = [v for v in monthly.values() if v is not None]
                    last_val = round(sum(valid) / len(valid), 1) if valid else None
                    _dc(ws, r, 2, sub_label, bg=bg)
                    for mi, m in enumerate(range(1, 13), 3):
                        _dc(ws, r, mi, monthly.get(m), bg=bg, nf='0.0')
                    _dc(ws, r, 15, last_val, bg=bg, bold=True, nf='0.0')
                    r += 1

            # 일조시간 블록
            if 'sunshine' in sdf.columns:
                bg = C['white']
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
                c = ws.cell(row=r, column=1, value='일조시간(hr)')
                c.font = Font(name=FONT, bold=True, size=9, color='000000')
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = _thin()
                ws.cell(row=r, column=2).border = _thin()
                ws.cell(row=r, column=2).fill = PatternFill(
                    start_color=bg, end_color=bg, fill_type='solid')
                monthly = _monthly_vals('sunshine', 'mean')
                valid = [v for v in monthly.values() if v is not None]
                annual_total = round(sum(valid), 1) if valid else None
                for mi, m in enumerate(range(1, 13), 3):
                    _dc(ws, r, mi, monthly.get(m), bg=bg, nf='#,##0.0')
                _dc(ws, r, 15, annual_total, bg=bg, bold=True, nf='#,##0.0')
                r += 1

            # 바람 블록
            wind_max_col = ('wind_gust' if 'wind_gust' in sdf.columns
                            else ('wind_max' if 'wind_max' in sdf.columns else None))
            wind_defs = [('wind_speed', '평균풍속', 'mean')]
            if wind_max_col:
                wind_defs.append((wind_max_col, '최대풍속', 'max'))
            avail_wind = [(ck, sl, ag) for ck, sl, ag in wind_defs if ck in sdf.columns]

            if avail_wind:
                bg = C['light_gray']
                merge_end = r + len(avail_wind) - 1
                ws.merge_cells(start_row=r, start_column=1, end_row=merge_end, end_column=1)
                c = ws.cell(row=r, column=1, value='바람(m/s)')
                c.font = Font(name=FONT, bold=True, size=9, color='000000')
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                c.border = _thin()
                for merge_r in range(r + 1, merge_end + 1):
                    ws.cell(row=merge_r, column=1).border = _thin()
                    ws.cell(row=merge_r, column=1).fill = PatternFill(
                        start_color=bg, end_color=bg, fill_type='solid')

                for col_key, sub_label, agg in avail_wind:
                    monthly = _monthly_vals(col_key, agg)
                    valid = [v for v in monthly.values() if v is not None]
                    if agg == 'mean':
                        last_val = round(sum(valid) / len(valid), 1) if valid else None
                    else:
                        last_val = round(max(valid), 1) if valid else None
                    _dc(ws, r, 2, sub_label, bg=bg)
                    for mi, m in enumerate(range(1, 13), 3):
                        _dc(ws, r, mi, monthly.get(m), bg=bg, nf='0.0')
                    _dc(ws, r, 15, last_val, bg=bg, bold=True, nf='0.0')
                    r += 1

            r += 1
            _note(ws, r, 1, total_width,
                  f'※ 자료 : 기상자료개방포털({yr_min}~{yr_max})')
            r += 3

    # ──────────────────────────────────
    # 추가 시트 1: 원본 데이터2 (세로형 Long Format)
    # ──────────────────────────────────

    def _sheet_raw2(self, wb, df):
        ws = wb.create_sheet("원본 데이터2")
        ws.freeze_panes = 'A2'

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 7
        ws.column_dimensions['C'].width = 5
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 12

        headers = ['날짜', '연도', '월', '관측소명', '항목', 'Data']
        for ci, h in enumerate(headers, 1):
            _hc(ws, 1, ci, h, sz=9, bg=C['dark_blue'])

        elem_cols = [(k, lbl) for k, lbl in ELEMENT_LABELS.items() if k in df.columns]

        row_r = 2
        for _, rec in df.iterrows():
            date_val = rec.get('date', None)
            if date_val is not None and pd.notna(date_val):
                date_val = pd.to_datetime(date_val).to_pydatetime()
            else:
                date_val = None
            year_val = rec.get('year', None)
            month_val = rec.get('month', None)
            stn_val = rec.get('station_name', '')
            for col_key, label in elem_cols:
                raw_val = rec.get(col_key, None)
                val = _safe_val(raw_val)
                c_date = ws.cell(row=row_r, column=1, value=date_val)
                c_date.number_format = 'YYYY-MM-DD'
                ws.cell(row=row_r, column=2, value=year_val)
                ws.cell(row=row_r, column=3, value=month_val)
                ws.cell(row=row_r, column=4, value=stn_val)
                ws.cell(row=row_r, column=5, value=label)
                ws.cell(row=row_r, column=6, value=val)
                row_r += 1

        ws.auto_filter.ref = f"A1:F{max(row_r - 1, 1)}"
        self._raw2_total_rows = max(row_r - 1, 1)
        self._raw2_item_labels = [lbl for _, lbl in elem_cols]

    # ──────────────────────────────────
    # 추가 시트 2: 피벗작업 (연도×월 강수량 합계)
    # ──────────────────────────────────

    def _sheet_pivot_work(self, wb, df):
        """피벗작업 시트 생성 (빈 시트 - 피벗 테이블은 _inject_pivot_tables에서 주입)"""
        ws = wb.create_sheet("피벗작업")
        ws.sheet_view.showGridLines = True
        ws.column_dimensions['A'].width = 12
        for ci in range(2, 15):
            ws.column_dimensions[get_column_letter(ci)].width = 9

    # ──────────────────────────────────
    # 추가 시트 3: 피벗작업2 (날짜×월 전체 Data 합계)
    # ──────────────────────────────────

    def _sheet_pivot_work2(self, wb, df):
        """피벗작업2 시트 생성 (빈 시트 - 피벗 테이블은 _inject_pivot_tables에서 주입)"""
        ws = wb.create_sheet("피벗작업2")
        ws.sheet_view.showGridLines = True
        ws.column_dimensions['A'].width = 13
        for ci in range(2, 15):
            ws.column_dimensions[get_column_letter(ci)].width = 9

    # ──────────────────────────────────
    # 추가 시트 4: Box Plot 분석
    # ──────────────────────────────────

    def _sheet_boxplot(self, wb, df):
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import matplotlib.font_manager as fm

        # 한글 폰트 설정 (docx_generator.py 방식과 동일)
        _KR_FONTS = ['Malgun Gothic', 'NanumGothic', 'AppleGothic', 'DejaVu Sans']
        available_fonts = {f.name for f in fm.fontManager.ttflist}
        for fn in _KR_FONTS:
            if fn in available_fonts:
                matplotlib.rcParams['font.family'] = fn
                break
        matplotlib.rcParams['axes.unicode_minus'] = False

        ws = wb.create_sheet("📦 Box Plot 분석")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions['A'].width = 14

        # 타이틀
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        c = ws.cell(row=1, column=1, value="Box Plot 분석  —  기온·강수량 분포 (월별·연도별)")
        c.font = Font(name=FONT, bold=True, size=13, color=C['white'])
        c.fill = PatternFill(start_color=C['dark_blue'], end_color=C['dark_blue'], fill_type='solid')
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[1].height = 28

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
        c2 = ws.cell(row=2, column=1,
                     value="※ 다이아몬드(◆) = 평균 / 상자 = IQR(Q1~Q3) / 수염 = 1.5×IQR / 점 = 이상치")
        c2.font = Font(name=FONT, size=8, color=C['gray'])
        c2.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[2].height = 13

        def _boxplot_img(data_list, labels, title, ylabel, color='skyblue', figsize=(10, 5)):
            fig, ax = plt.subplots(figsize=figsize)
            clean_data = [np.array([v for v in d if v is not None and not np.isnan(v)])
                          for d in data_list]
            clean_data = [d for d in clean_data if len(d) > 0]
            clean_labels = [lb for d, lb in zip(
                [np.array([v for v in d if v is not None and not np.isnan(v)])
                 for d in data_list], labels) if len(d) > 0]
            if not clean_data:
                plt.close(fig)
                return None
            bp = ax.boxplot(clean_data, labels=clean_labels, patch_artist=True,
                            showfliers=True, showmeans=True,
                            meanprops={'marker': 'D', 'markerfacecolor': 'orange',
                                       'markeredgecolor': 'orange', 'markersize': 6})
            for patch in bp['boxes']:
                patch.set_facecolor(color)
                patch.set_alpha(0.7)
            ax.set_title(title, fontsize=12, fontweight='bold')
            ax.set_ylabel(ylabel, fontsize=10)
            ax.grid(axis='y', alpha=0.3)
            plt.tight_layout()
            buf = _io.BytesIO()
            fig.savefig(buf, format='png', dpi=120)
            plt.close(fig)
            buf.seek(0)
            return buf

        def _insert_chart(ws, img_buf, anchor_row):
            if img_buf is None:
                return
            try:
                xl_img = XLImage(img_buf)
                xl_img.width = 700
                xl_img.height = 350
                ws.add_image(xl_img, f'A{anchor_row}')
            except Exception:
                pass

        chart_row = 3
        CHART_HEIGHT_ROWS = 18  # 차트 1개당 점유 행 수

        months = list(range(1, 13))
        month_labels = [str(m) for m in months]
        years = sorted(df['year'].dropna().unique().astype(int).tolist())
        year_labels = [str(y) for y in years]

        # ── 차트 1: 평균기온 월별 분포 ────────────────
        if 'temp_avg' in df.columns:
            data_by_month = [df[df['month'] == m]['temp_avg'].dropna().tolist()
                             for m in months]
            buf = _boxplot_img(data_by_month, month_labels,
                               '【평균기온】 월별 분포', '평균기온 (℃)',
                               color='lightcoral', figsize=(12, 5))
            _insert_chart(ws, buf, chart_row)
            chart_row += CHART_HEIGHT_ROWS

        # ── 차트 2: 평균기온 연도별 분포 ──────────────
        if 'temp_avg' in df.columns:
            data_by_year = [df[df['year'] == y]['temp_avg'].dropna().tolist()
                            for y in years]
            buf = _boxplot_img(data_by_year, year_labels,
                               '【평균기온】 연도별 분포', '평균기온 (℃)',
                               color='salmon', figsize=(max(10, len(years) * 0.8 + 2), 5))
            _insert_chart(ws, buf, chart_row)
            chart_row += CHART_HEIGHT_ROWS

        # ── 차트 3: 최고·최저기온 월별 분포 비교 ──────
        if 'temp_max' in df.columns or 'temp_min' in df.columns:
            fig, axes = plt.subplots(1, 2, figsize=(14, 5))
            for ax, col_key, label, color in [
                ('temp_max', '최고기온 (℃)', 'tomato'),
                ('temp_min', '최저기온 (℃)', 'steelblue'),
            ]:
                pass  # placeholder
            plt.close(fig)

            # 최고기온
            if 'temp_max' in df.columns:
                data_max = [df[df['month'] == m]['temp_max'].dropna().tolist() for m in months]
                buf_max = _boxplot_img(data_max, month_labels,
                                       '【최고기온】 월별 분포', '최고기온 (℃)',
                                       color='tomato', figsize=(12, 5))
                _insert_chart(ws, buf_max, chart_row)
                chart_row += CHART_HEIGHT_ROWS

            # 최저기온
            if 'temp_min' in df.columns:
                data_min = [df[df['month'] == m]['temp_min'].dropna().tolist() for m in months]
                buf_min = _boxplot_img(data_min, month_labels,
                                       '【최저기온】 월별 분포', '최저기온 (℃)',
                                       color='steelblue', figsize=(12, 5))
                _insert_chart(ws, buf_min, chart_row)
                chart_row += CHART_HEIGHT_ROWS

        # ── 차트 4: 강수량 월별 분포 ──────────────────
        if 'precipitation' in df.columns:
            data_precip_m = [df[df['month'] == m]['precipitation'].dropna().tolist()
                             for m in months]
            buf = _boxplot_img(data_precip_m, month_labels,
                               '【강수량】 월별 분포', '강수량 (mm)',
                               color='skyblue', figsize=(12, 5))
            _insert_chart(ws, buf, chart_row)
            chart_row += CHART_HEIGHT_ROWS

        # ── 차트 5: 강수량 연도별 분포 (유강우일, >0mm) ─
        if 'precipitation' in df.columns:
            df_rain = df[df['precipitation'] > 0]
            data_precip_y = [df_rain[df_rain['year'] == y]['precipitation'].dropna().tolist()
                             for y in years]
            buf = _boxplot_img(data_precip_y, year_labels,
                               '【강수량】 연도별 분포 (유강우일 >0mm)', '강수량 (mm)',
                               color='cornflowerblue',
                               figsize=(max(10, len(years) * 0.8 + 2), 5))
            _insert_chart(ws, buf, chart_row)
            chart_row += CHART_HEIGHT_ROWS

        # ── 통계 요약 표 ──────────────────────────────
        stat_row = chart_row + 1
        ws.merge_cells(start_row=stat_row, start_column=1, end_row=stat_row, end_column=9)
        c = ws.cell(row=stat_row, column=1, value="통계 요약  —  월별 사분위수 (IQR) 표")
        c.font = Font(name=FONT, bold=True, size=11, color=C['white'])
        c.fill = PatternFill(start_color=C['mid_blue'], end_color=C['mid_blue'], fill_type='solid')
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[stat_row].height = 22
        stat_row += 1

        stat_headers = ['월', '최소', 'Q1(25%)', '중앙값', '평균', 'Q3(75%)', '최대', 'IQR']

        def _write_stat_block(ws, start_row, col_key, block_label):
            if col_key not in df.columns:
                return start_row
            # 블록 타이틀
            ws.merge_cells(start_row=start_row, start_column=1,
                           end_row=start_row, end_column=len(stat_headers))
            c = ws.cell(row=start_row, column=1, value=f"[ {block_label} ] 월별 통계")
            c.font = Font(name=FONT, bold=True, size=9, color=C['white'])
            c.fill = PatternFill(start_color=C['dark_blue'], end_color=C['dark_blue'], fill_type='solid')
            c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[start_row].height = 16
            start_row += 1
            # 헤더
            for ci, h in enumerate(stat_headers, 1):
                _hc(ws, start_row, ci, h, sz=9, bg=C['mid_blue'])
            start_row += 1
            # 데이터
            for m in range(1, 13):
                mdf = df[df['month'] == m][col_key].dropna()
                if len(mdf) == 0:
                    _dc(ws, start_row, 1, f'{m}월', bg=C['light_blue'])
                    for ci in range(2, len(stat_headers) + 1):
                        _dc(ws, start_row, ci, None)
                else:
                    q1 = round(float(mdf.quantile(0.25)), 2)
                    median = round(float(mdf.median()), 2)
                    mean = round(float(mdf.mean()), 2)
                    q3 = round(float(mdf.quantile(0.75)), 2)
                    iqr = round(q3 - q1, 2)
                    vals = [f'{m}월', round(float(mdf.min()), 2), q1, median, mean,
                            q3, round(float(mdf.max()), 2), iqr]
                    for ci, v in enumerate(vals, 1):
                        bg = C['light_blue'] if ci == 1 else C['white']
                        bold = ci == 1
                        _dc(ws, start_row, ci, v, bg=bg, bold=bold, nf='0.00')
                start_row += 1
            return start_row + 1

        stat_row = _write_stat_block(ws, stat_row, 'temp_avg', '평균기온(℃)')
        stat_row = _write_stat_block(ws, stat_row, 'precipitation', '강수량(mm)')

    # ──────────────────────────────────
    # 피벗 테이블 XML 직접 주입
    # ──────────────────────────────────

    def _inject_pivot_tables(self, xlsx_bytes: bytes, df) -> bytes:
        """
        완성된 XLSX 바이트에 피벗 테이블 XML을 직접 주입합니다.
        openpyxl의 불완전한 피벗 테이블 지원을 우회합니다.
        """
        import zipfile
        import io as _io
        import re

        # ── 데이터 수집 ──────────────────────────────────
        years = sorted(df['year'].dropna().unique().astype(int).tolist())
        months = list(range(1, 13))
        stations = (sorted(df['station_name'].dropna().unique().tolist())
                    if 'station_name' in df.columns else ['관측소'])
        item_labels = getattr(self, '_raw2_item_labels', list(ELEMENT_LABELS.values()))
        total_rows = getattr(self, '_raw2_total_rows', 100000)

        # 고유 날짜 (string 형식)
        if 'date' in df.columns:
            dates = sorted(df['date'].dropna().dt.strftime('%Y-%m-%d').unique().tolist())
        else:
            dates = []

        n_years = len(years)
        n_dates = len(dates)
        n_stations = len(stations)
        n_labels = len(item_labels)

        # ── XML 생성 헬퍼 ─────────────────────────────────
        def esc(s):
            """XML 특수문자 이스케이프"""
            return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;')

        def str_shared(vals):
            return ''.join(f'<s v="{esc(v)}"/>' for v in vals)

        def num_shared(vals):
            return ''.join(f'<n v="{v}"/>' for v in vals)

        def pf_items(n):
            """pivotField items: <item x="0"/><item x="1"/>...<item t="default"/>"""
            return ''.join(f'<item x="{i}"/>' for i in range(n)) + '<item t="default"/>'

        def row_col_items(n):
            """rowItems or colItems: <i><x v="0"/></i>...<i t="grand"><x/></i>"""
            return ''.join(f'<i><x v="{i}"/></i>' for i in range(n)) + '<i t="grand"><x/></i>'

        # ── pivotCacheDefinition1.xml ───────────────────────────────
        # r:id 제거 (records 파일 불필요, refreshOnLoad로 Excel이 재구성)
        # 날짜/Data sharedItems는 최소한으로 (pivot axis로 사용하지 않음)
        cache_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<pivotCacheDefinition xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" refreshOnLoad="1" createdVersion="5" refreshedVersion="5" minRefreshableVersion="3" recordCount="0">
<cacheSource type="worksheet"><worksheetSource ref="A1:F{total_rows}" sheet="원본 데이터2"/></cacheSource>
<cacheFields count="6">
<cacheField name="날짜" numFmtId="0"><sharedItems/></cacheField>
<cacheField name="연도" numFmtId="0"><sharedItems containsString="0" containsBlank="0" containsNumber="1" containsInteger="1" minValue="{years[0] if years else 0}" maxValue="{years[-1] if years else 0}" count="{n_years}">{num_shared(years)}</sharedItems></cacheField>
<cacheField name="월" numFmtId="0"><sharedItems containsString="0" containsBlank="0" containsNumber="1" containsInteger="1" minValue="1" maxValue="12" count="12">{num_shared(months)}</sharedItems></cacheField>
<cacheField name="관측소명" numFmtId="0"><sharedItems count="{n_stations}">{str_shared(stations)}</sharedItems></cacheField>
<cacheField name="항목" numFmtId="0"><sharedItems count="{n_labels}">{str_shared(item_labels)}</sharedItems></cacheField>
<cacheField name="Data" numFmtId="0"><sharedItems containsString="0" containsBlank="1" containsMixedTypes="1"/></cacheField>
</cacheFields>
</pivotCacheDefinition>'''

        # ── pivotTable1.xml (피벗작업: 연도×월) ─────────────────────
        # r:id 제거 (pivotTableDefinition에는 유효하지 않음)
        # Data 필드: axis="axisValues" → dataField="1" (OOXML 표준)
        pivot1_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<pivotTableDefinition xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" name="PivotTable_강수량" cacheId="0" applyNumberFormats="0" applyBorderFormats="0" applyFontFormats="0" applyPatternFormats="0" applyAlignmentFormats="0" applyWidthHeightFormats="1" dataCaption="값" updatedVersion="5" minRefreshableVersion="3" createdVersion="5" useAutoFormatting="1" itemPrintTitles="1" indent="0" compact="0" compactData="0" multipleFieldFilters="0">
<location ref="A3:N{3+n_years}" firstHeaderRow="1" firstDataRow="2" firstDataCol="1"/>
<pivotFields count="6">
<pivotField compact="0" outline="0" subtotalTop="0" showAll="0"/>
<pivotField axis="axisRow" compact="0" outline="0" subtotalTop="0" showAll="0"><items count="{n_years+1}">{pf_items(n_years)}</items></pivotField>
<pivotField axis="axisCol" compact="0" outline="0" subtotalTop="0" showAll="0"><items count="13">{pf_items(12)}</items></pivotField>
<pivotField compact="0" outline="0" subtotalTop="0" showAll="0"/>
<pivotField compact="0" outline="0" subtotalTop="0" showAll="0"/>
<pivotField dataField="1" compact="0" outline="0" subtotalTop="0" showAll="0"/>
</pivotFields>
<rowFields count="1"><field x="1"/></rowFields>
<rowItems count="{n_years+1}">{row_col_items(n_years)}</rowItems>
<colFields count="1"><field x="2"/></colFields>
<colItems count="13">{row_col_items(12)}</colItems>
<dataFields count="1"><dataField name="합계 : Data" fld="5" baseField="1" baseItem="0"/></dataFields>
</pivotTableDefinition>'''

        # ── pivotTable2.xml (피벗작업2: 항목×연도) ─────────────────────
        n_row2 = max(n_labels, 1)
        pivot2_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<pivotTableDefinition xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" name="PivotTable_항목별" cacheId="0" applyNumberFormats="0" applyBorderFormats="0" applyFontFormats="0" applyPatternFormats="0" applyAlignmentFormats="0" applyWidthHeightFormats="1" dataCaption="값" updatedVersion="5" minRefreshableVersion="3" createdVersion="5" useAutoFormatting="1" itemPrintTitles="1" indent="0" compact="0" compactData="0" multipleFieldFilters="0">
<location ref="A3:{get_column_letter(1+n_years+1)}{3+n_row2}" firstHeaderRow="1" firstDataRow="2" firstDataCol="1"/>
<pivotFields count="6">
<pivotField compact="0" outline="0" subtotalTop="0" showAll="0"/>
<pivotField axis="axisCol" compact="0" outline="0" subtotalTop="0" showAll="0"><items count="{n_years+1}">{pf_items(n_years)}</items></pivotField>
<pivotField compact="0" outline="0" subtotalTop="0" showAll="0"/>
<pivotField compact="0" outline="0" subtotalTop="0" showAll="0"/>
<pivotField axis="axisRow" compact="0" outline="0" subtotalTop="0" showAll="0"><items count="{n_row2+1}">{pf_items(n_row2)}</items></pivotField>
<pivotField dataField="1" compact="0" outline="0" subtotalTop="0" showAll="0"/>
</pivotFields>
<rowFields count="1"><field x="4"/></rowFields>
<rowItems count="{n_row2+1}">{row_col_items(n_row2)}</rowItems>
<colFields count="1"><field x="1"/></colFields>
<colItems count="{n_years+1}">{row_col_items(n_years)}</colItems>
<dataFields count="1"><dataField name="합계 : Data" fld="5" baseField="4" baseItem="0"/></dataFields>
</pivotTableDefinition>'''

        # ── Relationship XMLs ─────────────────────────────────────────
        CACHE_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition"
        REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"

        # 캐시 rels: r:id 없으므로 빈 관계 (records 불필요)
        cache_rels_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<Relationships xmlns="{REL_NS}"/>'
        )

        pt1_rels_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="{REL_NS}"><Relationship Id="rId1" Type="{CACHE_NS}" Target="../pivotCache/pivotCacheDefinition1.xml"/></Relationships>'''

        pt2_rels_xml = pt1_rels_xml  # 같은 캐시 참조

        # ── ZIP 처리 ──────────────────────────────────────────────────
        in_buf = _io.BytesIO(xlsx_bytes)
        out_buf = _io.BytesIO()

        with zipfile.ZipFile(in_buf, 'r') as zin:
            # workbook.xml 및 관계 파일 읽기
            wb_rels = zin.read('xl/_rels/workbook.xml.rels').decode('utf-8')
            wb_xml = zin.read('xl/workbook.xml').decode('utf-8')

            # 시트 이름 → r:id 매핑 (속성 순서 무관하게 태그 전체에서 추출)
            sheet_ridmap = {}
            for tag_m in re.finditer(r'<sheet\s[^>]+>', wb_xml):
                tag = tag_m.group(0)
                name_m = re.search(r'name="([^"]+)"', tag)
                rid_m = re.search(r'r:id="(rId\d+)"', tag)
                if name_m and rid_m:
                    sheet_ridmap[name_m.group(1)] = rid_m.group(1)

            # r:id → 파일 경로 매핑 (속성 순서 무관하게 태그 전체에서 추출)
            rid_to_file = {}
            for tag_m in re.finditer(r'<Relationship[^>]+>', wb_rels):
                tag = tag_m.group(0)
                id_m = re.search(r'Id="(rId\d+)"', tag)
                tgt_m = re.search(r'Target="([^"]+)"', tag)
                if id_m and tgt_m:
                    rid_to_file[id_m.group(1)] = tgt_m.group(1)

            def sheet_file_num(sheet_name):
                rid = sheet_ridmap.get(sheet_name, '')
                path = rid_to_file.get(rid, '')
                m = re.search(r'sheet(\d+)\.xml', path)
                return int(m.group(1)) if m else None

            sn1 = sheet_file_num("피벗작업")
            sn2 = sheet_file_num("피벗작업2")
            if sn1 is None or sn2 is None:
                return xlsx_bytes  # 시트를 찾을 수 없으면 원본 반환

            sheet1_rels_path = f'xl/worksheets/_rels/sheet{sn1}.xml.rels'
            sheet2_rels_path = f'xl/worksheets/_rels/sheet{sn2}.xml.rels'

            # workbook.xml에 pivotCaches 섹션 추가
            new_wb_xml = wb_xml
            if '<pivotCaches>' not in wb_xml:
                pivot_caches_xml = (
                    '<pivotCaches>'
                    '<pivotCache xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
                    ' cacheId="0" r:id="rIdPivot1"/>'
                    '</pivotCaches>'
                )
                # OOXML 스키마 순서: pivotCaches는 calcPr 뒤에 와야 함
                # calcPr은 보통 self-closing (<calcPr .../>)
                calcpr_m = re.search(r'<calcPr[^/]*/>', wb_xml)
                if calcpr_m:
                    ins_pos = calcpr_m.end()
                    new_wb_xml = wb_xml[:ins_pos] + pivot_caches_xml + wb_xml[ins_pos:]
                elif '</calcPr>' in wb_xml:
                    new_wb_xml = wb_xml.replace('</calcPr>', '</calcPr>' + pivot_caches_xml, 1)
                else:
                    new_wb_xml = wb_xml.replace('</workbook>', pivot_caches_xml + '</workbook>', 1)

            # workbook.xml.rels에 캐시 관계 추가 (상대경로 사용)
            new_wb_rels = wb_rels
            if 'pivotCacheDefinition' not in wb_rels:
                cache_rel = (
                    '<Relationship Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition"'
                    ' Target="pivotCache/pivotCacheDefinition1.xml" Id="rIdPivot1"/>'
                )
                new_wb_rels = wb_rels.replace('</Relationships>', cache_rel + '</Relationships>')

            # [Content_Types].xml에 피벗 관련 content type 추가
            ct_xml = zin.read('[Content_Types].xml').decode('utf-8')
            new_ct_xml = ct_xml
            CT_PIVOT_CACHE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheDefinition+xml'
            CT_PIVOT_TABLE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.pivotTable+xml'

            additions = ''
            if 'pivotCacheDefinition' not in ct_xml:
                additions += f'<Override PartName="/xl/pivotCache/pivotCacheDefinition1.xml" ContentType="{CT_PIVOT_CACHE}"/>'
            if CT_PIVOT_TABLE not in ct_xml:
                additions += f'<Override PartName="/xl/pivotTables/pivotTable1.xml" ContentType="{CT_PIVOT_TABLE}"/>'
                additions += f'<Override PartName="/xl/pivotTables/pivotTable2.xml" ContentType="{CT_PIVOT_TABLE}"/>'
            if additions:
                new_ct_xml = ct_xml.replace('</Types>', additions + '</Types>')

            # 피벗 테이블 관계 항목
            pt_rel_entry = (
                '<Relationship Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable"'
                ' Target="../pivotTables/pivotTable1.xml" Id="rIdPT1"/>'
            )
            pt2_rel_entry = (
                '<Relationship Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable"'
                ' Target="../pivotTables/pivotTable2.xml" Id="rIdPT1"/>'
            )

            skip_paths = {
                'xl/workbook.xml',
                'xl/_rels/workbook.xml.rels',
                '[Content_Types].xml',
                sheet1_rels_path,
                sheet2_rels_path,
                'xl/pivotCache/pivotCacheDefinition1.xml',
                'xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels',
                'xl/pivotTables/pivotTable1.xml',
                'xl/pivotTables/pivotTable2.xml',
                'xl/pivotTables/_rels/pivotTable1.xml.rels',
                'xl/pivotTables/_rels/pivotTable2.xml.rels',
            }

            existing_names = {item.filename for item in zin.infolist()}

            with zipfile.ZipFile(out_buf, 'w', zipfile.ZIP_DEFLATED) as zout:
                # 기존 파일 복사 (수정 대상 제외)
                for item in zin.infolist():
                    if item.filename in skip_paths:
                        continue
                    zout.writestr(item, zin.read(item.filename))

                # 수정된 파일 쓰기
                zout.writestr('xl/workbook.xml', new_wb_xml.encode('utf-8'))
                zout.writestr('xl/_rels/workbook.xml.rels', new_wb_rels.encode('utf-8'))
                zout.writestr('[Content_Types].xml', new_ct_xml.encode('utf-8'))

                # 피벗 캐시 (records 없음 — refreshOnLoad가 Excel에서 재구성)
                zout.writestr('xl/pivotCache/pivotCacheDefinition1.xml', cache_xml.encode('utf-8'))
                zout.writestr('xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels',
                              cache_rels_xml.encode('utf-8'))

                # 피벗 테이블 XML
                zout.writestr('xl/pivotTables/pivotTable1.xml', pivot1_xml.encode('utf-8'))
                zout.writestr('xl/pivotTables/pivotTable2.xml', pivot2_xml.encode('utf-8'))
                zout.writestr('xl/pivotTables/_rels/pivotTable1.xml.rels', pt1_rels_xml.encode('utf-8'))
                zout.writestr('xl/pivotTables/_rels/pivotTable2.xml.rels', pt2_rels_xml.encode('utf-8'))

                # 시트 rels 파일 (피벗 테이블 링크 추가)
                for rels_path, rel_entry in [
                    (sheet1_rels_path, pt_rel_entry),
                    (sheet2_rels_path, pt2_rel_entry),
                ]:
                    if rels_path in existing_names:
                        existing = zin.read(rels_path).decode('utf-8')
                        new_rels = existing.replace('</Relationships>', rel_entry + '</Relationships>')
                    else:
                        new_rels = (
                            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                            f'<Relationships xmlns="{REL_NS}">{rel_entry}</Relationships>'
                        )
                    zout.writestr(rels_path, new_rels.encode('utf-8'))

        return out_buf.getvalue()
